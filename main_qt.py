import matplotlib
matplotlib.use('TkAgg')

import numpy as np
import matplotlib.pyplot as plt
from os.path import splitext, basename
from time import perf_counter
from matplotlib.patches import Rectangle, Circle
from matplotlib.widgets import AxesWidget, Button, TextBox, Slider
from scipy.interpolate import splprep, splev
from scipy.ndimage import uniform_filter1d
from tkinter import filedialog, messagebox

# 可选依赖：openpyxl 用于 Excel 读写（CSV 走 numpy）
try:
    from openpyxl import load_workbook, Workbook
    HAS_EXCEL = True
except ImportError:
    HAS_EXCEL = False
    print("提示: 安装 openpyxl 可支持 Excel 格式 (pip install openpyxl)")

# 可选依赖：tkinterdnd2 用于文件拖拽
try:
    from tkinterdnd2 import TkinterDnD, DND_FILES
    HAS_DND = True
except ImportError:
    HAS_DND = False
    print("提示: 安装 tkinterdnd2 可启用文件拖拽功能 (pip install tkinterdnd2)")

# ===================== 核心修复：精准、安全的 Matplotlib 补丁 =====================
def patch_matplotlib_widgets_safely():
    """
    精准修复 Matplotlib 组件基类的全局事件分发 Bug，增加防重复修改锁，彻底避免无限递归
    """
    base_class = AxesWidget
    
    # 检查是否已经修补过，防止重复执行
    if not getattr(base_class, '_is_patched_for_resize', False):
        orig_connect = base_class.connect_event
        
        def safe_connect(self, event, callback):
            def safe_callback(evt):
                if evt is None or not hasattr(evt, 'inaxes'):
                    return 
                return callback(evt)
            return orig_connect(self, event, safe_callback)
        
        base_class.connect_event = safe_connect
        base_class._is_patched_for_resize = True

# 启动补丁
patch_matplotlib_widgets_safely()


class HarmoniousCurvesEditor:
    def __init__(self, ax, fig, ax_n_input, ax_radio=None, color_list=None):
        self.fig = fig
        self.ax = ax
        self.canvas = ax.figure.canvas
        self.curves = []

        # 曲线激活锁定（单击选中 + 双击锁定，最终活跃 = 两者并集）
        self.active_curve_indices = set()   # 单击选中的（单个）
        self.locked_curve_indices = set()   # 双击锁定的（多个，独立保留）
        self.pinned_curve_indices = set()   # Pin 按钮固定：始终高亮，永不被选中/拖拽
        self.current_label = "All"

        # 单点拖拽
        self._active_curve_idx = None
        self._active_point_idx = None

        # 增大容差像素，让鼠标更容易选中
        self._epsilon = 30

        self.neighbor_num = 0
        self.ax_n_input = ax_n_input
        self._typing_axes = {ax_n_input}  # 在这些区域内不响应 delete/backspace

        # 框选 & 批量移动
        self._selecting = False
        self._select_rect = None
        self._select_start = None
        self._selected_points = []
        self._batch_moving = False
        self._batch_origin = None
        self._selection_x_bounds = None

        # Shift+框选 → 可调窗口模式
        self._adjust_mode = False
        self._adjust_rect = None        # Rectangle patch
        self._adjust_handles = []       # 8 个 Rectangle 手柄
        self._adjust_data = None        # {x0, y0, x1, y1}
        self._adjust_drag_edge = None   # 'left'|'right'|'top'|'bottom'|'center'|'tl'|'tr'|'bl'|'br'
        self._adjust_drag_start = None
        self._shift_selecting = False
        self._handle_radius = 8         # 手柄命中半径（像素）
        self._last_motion_time = 0      # motion 节流计时器（秒）
        self._clipboard = None          # Ctrl+C/V: {'curve_idx': int, 'y_values': [...]}

        # 性能：motion 事件节流
        self._last_motion_time = 0      # 上次处理 motion 的时间（秒）
        self._motion_interval = 0.030   # 节流间隔 30ms ≈ 33fps

        # Undo 栈 & 临时操作状态备份
        self.undo_stack = []
        self._in_continuous_drag = False
        self._slider_base_state = None
        self._is_sliding = False  # 标记当前是否正在拖动滑杆
        self._noise_seed = None   # 单次拖拽缓存噪点种子，松手后重置
        self._tab_index = 0       # Tab 切换曲线索引，默认 All
        self._last_npts_val = None  # 上次同步到输入框的点数，避免覆盖用户输入

        # 组件引用初始化
        self.radio_menu = None
        self.radio_labels = []
        self.slider_scale = None
        self.slider_smooth = None
        self.slider_noise = None
        self._resample_box = None  # 外部注入的重采样输入框引用

        # 文件 I/O 状态
        self.current_file_path = None      # 当前文件路径（用于保存时推断格式）
        self.color_list = color_list or [] # 15 色调色板
        self.ax_radio = ax_radio           # 右侧单选面板 axes

        # 拖拽提示 & 单选面板顶层元素引用
        self.drop_hint = None              # 空数据时的拖拽提示文字
        self._top_dots = None              # 面板顶层圆点 scatter
        self._radio_circles = []           # 面板 Circle patch 列表
        self._color_patches = []           # 面板色块 Rectangle 列表
        self._lock_icons = []              # 面板锁定图标 Text 列表
  
        # 底部状态信息 — 大号数字加粗，描述浅色跟随
        self.info_num = self.fig.text(
            0.05, 0.02, "", fontsize=15, fontweight='bold',
            color='#1a1e24', va='bottom', ha='left',
        )
        self.info_tag = self.fig.text(
            0.15, 0.02, "", fontsize=10, color='#4a5568',
            fontweight='normal', va='bottom', ha='left',
        )

        # 事件绑定
        self.canvas.mpl_connect('button_press_event', self.on_press)
        self.canvas.mpl_connect('motion_notify_event', self.on_motion)
        self.canvas.mpl_connect('button_release_event', self.on_release)
        self.canvas.mpl_connect('key_press_event', self.on_key)
        self.canvas.mpl_connect('button_press_event', self.on_double_click)

        # Tk 级别绑定 Tab（matplotlib key_press_event 收不到 Tab）
        self._setup_tk_tab()
        # 点击绘图区时确保键盘焦点回到 canvas
        self.canvas.mpl_connect('button_press_event', self._ensure_focus)

        self._update_selection_info()

    def set_radio_ref(self, radio_menu, labels):
        self.radio_menu = radio_menu
        self.radio_labels = labels

    def _is_toolbar_active(self):
        toolbar = self.fig.canvas.manager.toolbar
        if toolbar is not None:
            mode = toolbar.mode.lower()
            if 'zoom' in mode or 'pan' in mode:
                return True
        return False

    def on_double_click(self, event):
        if not hasattr(event, 'inaxes') or event.inaxes is None:
            return
        if event.button != 1 or not event.dblclick:
            return
        if event.inaxes != self.ax or self._is_toolbar_active():
            return
        c_idx, _ = self._get_closest_point(event)
        if c_idx is None:
            return
        target_label = f"Curve {c_idx}"
        if target_label in self.radio_labels:
            self.toggle_active_curve(target_label)

    def _get_effective_active(self):
        """返回当前活跃的曲线集合 = 单击选中 ∪ 双击锁定"""
        return self.active_curve_indices | self.locked_curve_indices

    def set_active_curve(self, label):
        """单击：设置单选曲线（替换之前单击的，不影响锁定的）。All = 清空单击选中"""
        if label == "All":
            self.active_curve_indices = set()
        else:
            for idx, curve in enumerate(self.curves):
                if curve['name'] == label:
                    self.active_curve_indices = {idx}
                    break
        self._update_current_label()
        effective = self._get_effective_active()
        if self._selected_points and len(effective) > 0:
            self._selected_points = [(c, p) for c, p in self._selected_points if c in effective]
        self._refresh_radio_panel()
        self._refresh_visual_style()
        self._reset_slider_widgets()
        self._update_selection_info()

    def toggle_active_curve(self, label):
        """双击：锁定/解锁曲线（toggle，独立保留，不影响单击选中）。All = 清空所有锁定"""
        if label == "All":
            self.locked_curve_indices = set()
        else:
            for idx, curve in enumerate(self.curves):
                if curve['name'] == label:
                    if idx in self.locked_curve_indices:
                        self.locked_curve_indices.discard(idx)
                    else:
                        self.locked_curve_indices.add(idx)
                    break
        self._update_current_label()
        effective = self._get_effective_active()
        if self._selected_points and len(effective) > 0:
            self._selected_points = [(c, p) for c, p in self._selected_points if c in effective]
        self._refresh_radio_panel()
        self._refresh_visual_style()
        self._reset_slider_widgets()
        self._update_selection_info()

    def toggle_pin(self, event=None):
        """Pin 按钮：将当前活跃曲线固定——始终高亮，但永不被选中/拖拽"""
        effective = self._get_effective_active()
        if len(effective) == 0:
            # 无活跃曲线 → 对所有曲线切换 pin（全 pin 则取消全 pin）
            if len(self.pinned_curve_indices) == len(self.curves):
                self.pinned_curve_indices.clear()
                print("  📌 取消全部固定")
            else:
                self.pinned_curve_indices = set(range(len(self.curves)))
                print(f"  📌 固定全部 {len(self.curves)} 条曲线")
        else:
            for idx in effective:
                if idx in self.pinned_curve_indices:
                    self.pinned_curve_indices.discard(idx)
                    print(f"  📌 取消固定 {self.curves[idx]['name']}")
                else:
                    self.pinned_curve_indices.add(idx)
                    print(f"  📌 固定 {self.curves[idx]['name']}")
        self._refresh_visual_style()
        self._refresh_radio_panel()
        self._update_selection_info()

    def _update_current_label(self):
        effective = self._get_effective_active()
        if len(effective) == 0:
            self.current_label = "All"
        elif len(effective) == 1:
            self.current_label = self.curves[next(iter(effective))]['name']
        else:
            self.current_label = f"{len(effective)} curves"

    def _refresh_visual_style(self):
        effective = self._get_effective_active()
        # 预计算：一次遍历分组，避免每曲线重复扫描 _selected_points
        sel_by_curve = {}
        for c, p in self._selected_points:
            sel_by_curve.setdefault(c, []).append(p)

        for idx, curve in enumerate(self.curves):
            is_curve_active = (len(effective) == 0 or idx in effective
                               or idx in self.pinned_curve_indices)
            if self._selection_x_bounds is not None:
                curve['ctrl_points'].set_alpha(0.06)
                curve['spline_line'].set_alpha(0.18)
                curve['spline_line'].set_linewidth(1.2)
                sel_p_idxs = sel_by_curve.get(idx)
                if is_curve_active and sel_p_idxs:
                    xs_p = curve['x'][sel_p_idxs]
                    ys_p = curve['y'][sel_p_idxs]
                    curve['hl_ctrl_points'].set_data(xs_p, ys_p)
                    curve['hl_ctrl_points'].set_alpha(1.0)
                    curve['hl_ctrl_points'].set_markersize(7)
                    x_fine, y_fine = curve['fine_x_full'], curve['fine_y_full']
                    xmin, xmax = self._selection_x_bounds
                    mask = (x_fine >= xmin) & (x_fine <= xmax)
                    if np.any(mask):
                        curve['hl_spline_line'].set_data(x_fine[mask], y_fine[mask])
                        curve['hl_spline_line'].set_alpha(1.0)
                        curve['hl_spline_line'].set_linewidth(4.0)
                    else:
                        curve['hl_spline_line'].set_data([], [])
                else:
                    curve['hl_spline_line'].set_data([], [])
                    curve['hl_ctrl_points'].set_data([], [])
            else:
                curve['hl_spline_line'].set_data([], [])
                curve['hl_ctrl_points'].set_data([], [])
                if is_curve_active:
                    is_all = len(effective) == 0
                    alpha_ctrl = 0.45 if is_all else 0.85
                    lw = 2.5 if is_all else 3.8
                    ms = 5 if is_all else 7
                    curve['ctrl_points'].set_alpha(alpha_ctrl)
                    curve['ctrl_points'].set_markersize(ms)
                    curve['spline_line'].set_alpha(1.0)
                    curve['spline_line'].set_linewidth(lw)
                else:
                    curve['ctrl_points'].set_alpha(0.10)
                    curve['ctrl_points'].set_markersize(4)
                    curve['spline_line'].set_alpha(0.22)
                    curve['spline_line'].set_linewidth(1.4)
        self.canvas.draw_idle()

    def _update_selection_info(self):
        """显示当前选中的控制点信息，并实时初始化 N-pts 右侧重采样输入框"""
        if not self.curves:
            self.info_num.set_text("")
            self.info_tag.set_text("")
            return

        current_resample_val = ""

        if self._selected_points:
            # 1. 处于框选状态：统计每条曲线被选中的点数
            by_curve = {}
            for c, p in self._selected_points:
                by_curve.setdefault(c, []).append(p)
            
            n = len(self._selected_points)
            detail = ", ".join(f"C{c}: {len(p_)}p" for c, p_ in sorted(by_curve.items()))
            self.info_num.set_text(f"{n} pts")
            self.info_tag.set_text(f"Box-sel  ({detail})")

            # 提取所有被选中曲线片段的点数集合
            selected_counts = {len(p_idxs) for p_idxs in by_curve.values()}
            if len(selected_counts) == 1:
                # 只有当框选的每条曲线片段内的点数完全一致时，才在输入框初始化该数字
                current_resample_val = str(selected_counts.pop())
            else:
                current_resample_val = "---"  # 不一致则显示占位符，提示输入

        elif len(self._get_effective_active()) > 0:
            # 2. 未框选，但选中了若干条曲线
            active_indices = sorted(self._get_effective_active())
            total = sum(len(self.curves[i]['y']) for i in active_indices)
            self.info_num.set_text(f"{total} pts")
            self.info_tag.set_text(f"{self.current_label}")
            counts = {len(self.curves[i]['y']) for i in active_indices}
            current_resample_val = str(counts.pop()) if len(counts) == 1 else "---"
        else:
            # 3. 默认 All 状态，未进行框选
            total = sum(len(c['y']) for c in self.curves)
            self.info_num.set_text(f"{total} pts")
            nc = len(self.curves)
            all_counts = {len(c['y']) for c in self.curves}
            if len(all_counts) == 1:
                self.info_tag.set_text(f"All ({nc} curves x {all_counts.pop()})")
            else:
                self.info_tag.set_text(f"All ({nc} curves, ragged)")
            current_resample_val = str(all_counts.pop()) if len(all_counts) == 1 else "---"

        # 同步输入框：点数一致填数字，不一致填 "---"
        if hasattr(self, '_resample_box') and self._resample_box:
            if current_resample_val != self._last_npts_val:
                self._resample_box.set_val(current_resample_val)
                self._last_npts_val = current_resample_val

        self.canvas.draw_idle()

    def _canvas_has_focus(self):
        try:
            canvas = self.fig.canvas.get_tk_widget()
            focused = canvas.focus_get()
            return focused is None or focused is canvas
        except Exception:
            return True

    def _setup_tk_tab(self):
        try:
            canvas_widget = self.fig.canvas.get_tk_widget()
            def _on_tk_tab(event):
                if not self.radio_labels or len(self.curves) == 0:
                    return
                # Tab 在 All → Curve 0 → Curve 1 → ... 之间单选切换
                self._tab_index = (self._tab_index + 1) % (len(self.curves) + 1)
                if self._tab_index == 0:
                    self.set_active_curve("All")
                else:
                    self.set_active_curve(self.curves[self._tab_index - 1]['name'])
                return 'break'
            canvas_widget.bind('<Tab>', _on_tk_tab)
        except Exception:
            pass

    def _ensure_focus(self, event=None):
        try:
            self.fig.canvas.get_tk_widget().focus_set()
        except Exception:
            pass

    def _clear_selection(self):
        self._selected_points.clear()
        self._batch_moving = False
        self._selection_x_bounds = None
        if self._select_rect is not None:
            self._select_rect.remove()
            self._select_rect = None
        self._exit_adjust_mode()
        self._refresh_visual_style()
        self._reset_slider_widgets()
        self._update_selection_info()

    # ===================== Shift+框选：可调窗口模式 =====================

    def _enter_adjust_mode(self, x0, y0, x1, y1):
        """进入可调窗口模式：显示橙色虚框 + 8 个拖拽手柄"""
        self._adjust_mode = True
        self._adjust_data = {
            'x0': min(x0, x1), 'y0': min(y0, y1),
            'x1': max(x0, x1), 'y1': max(y0, y1),
        }
        self._adjust_drag_edge = None
        self._adjust_drag_start = None

        # 橙色虚线矩形
        d = self._adjust_data
        self._adjust_rect = Rectangle(
            (d['x0'], d['y0']), d['x1'] - d['x0'], d['y1'] - d['y0'],
            fill=True, facecolor='#FF6B35', alpha=0.10,
            edgecolor='#FF6B35', linestyle=(0, (5, 3)), linewidth=2.2,
            zorder=50,
        )
        self.ax.add_patch(self._adjust_rect)

        # 8 个白色小方手柄
        self._build_adjust_handles()
        self._update_adjust_selection_preview()
        self.canvas.draw_idle()

    def _exit_adjust_mode(self):
        """退出可调窗口模式，清理所有 adjust 图元"""
        self._adjust_mode = False
        self._adjust_drag_edge = None
        self._adjust_drag_start = None
        self._shift_selecting = False
        if self._adjust_rect is not None:
            self._adjust_rect.remove()
            self._adjust_rect = None
        for h in self._adjust_handles:
            h.remove()
        self._adjust_handles.clear()
        self._adjust_data = None

    def _confirm_adjust(self):
        """确认当前可调窗口 → 固化为最终框选"""
        if self._adjust_data is None:
            return
        d = self._adjust_data
        self._selection_x_bounds = (d['x0'], d['x1'])
        self._selected_points = self._select_points_in_rect(
            d['x0'], d['y0'], d['x1'], d['y1'])
        self._exit_adjust_mode()
        self._refresh_visual_style()
        self._update_selection_info()
        print(f"  ✓ 窗口确认 — 选中 {len(self._selected_points)} 个控制点")

    def _cancel_adjust(self):
        """取消可调窗口"""
        self._exit_adjust_mode()
        self._selection_x_bounds = None
        self._selected_points.clear()
        self._refresh_visual_style()
        self._update_selection_info()
        print("  ✗ 窗口取消")

    def _build_adjust_handles(self):
        """在 adjust 矩形四角 + 四边中点创建手柄"""
        for h in self._adjust_handles:
            h.remove()
        self._adjust_handles.clear()

        d = self._adjust_data
        x0, y0, x1, y1 = d['x0'], d['y0'], d['x1'], d['y1']
        xm, ym = (x0 + x1) / 2, (y0 + y1) / 2

        hw, hh = self._handle_size_data()
        positions = {
            'bl': (x0, y0), 'bm': (xm, y0), 'br': (x1, y0),
            'ml': (x0, ym), 'mr': (x1, ym),
            'tl': (x0, y1), 'tm': (xm, y1), 'tr': (x1, y1),
        }
        for key, (hx, hy) in positions.items():
            rect = Rectangle(
                (hx - hw / 2, hy - hh / 2), hw, hh,
                facecolor='white', edgecolor='#FF6B35',
                linewidth=2.0, zorder=55,
            )
            self.ax.add_patch(rect)
            self._adjust_handles.append(rect)

    def _handle_size_data(self):
        """返回手柄在数据坐标中的半宽/半高（约 _handle_radius 像素）"""
        bbox = self.ax.get_window_extent()
        if bbox is None:
            return 0.02, 0.02
        px_w, px_h = bbox.width, bbox.height
        xl = self.ax.get_xlim()
        yl = self.ax.get_ylim()
        w_data = max(xl[1] - xl[0], 1e-9)
        h_data = max(yl[1] - yl[0], 1e-9)
        return (self._handle_radius * w_data / px_w,
                self._handle_radius * h_data / px_h)

    def _get_data_coords(self, event):
        """获取主图数据坐标。主图内用 event 坐标（精确），跨 axes 用 figure 回退"""
        # 鼠标在主图 axes 内 → 直接取，精度最高
        if hasattr(event, 'inaxes') and event.inaxes == self.ax:
            if event.xdata is not None and event.ydata is not None:
                return event.xdata, event.ydata
        
        # 鼠标跨到滑杆/面板/外部 → 从全局像素坐标（event.x, event.y）直接反推
        if event.x is not None and event.y is not None:
            try:
                # 传入真实的像素点 (event.x, event.y)，直接逆变换到数据坐标
                return self.ax.transData.inverted().transform([(event.x, event.y)])[0]
            except Exception:
                pass
                
        return event.xdata, event.ydata

    def _handle_adjust_drag(self, event):
        """根据 _adjust_drag_edge 拖拽调整窗口大小或移动（30ms 节流，支持出界）"""
        if self._adjust_data is None or self._adjust_drag_start is None:
            return

        now = perf_counter()
        if now - self._last_motion_time < 0.030:
            return
        self._last_motion_time = now

        xd, yd = self._get_data_coords(event)
        if xd is None or yd is None:
            return

        d = self._adjust_data
        dx = xd - self._adjust_drag_start[0]
        dy = yd - self._adjust_drag_start[1]
        edge = self._adjust_drag_edge
        x0, y0, x1, y1 = d['x0'], d['y0'], d['x1'], d['y1']

        # 用 np.nextafter 保证 y0<y1, x0<x1 永不翻转 — 零魔数，适配任意量级
        if edge == 'center':
            d['x0'] += dx; d['x1'] += dx
            d['y0'] += dy; d['y1'] += dy
        elif 'l' in edge or edge in ('left',):
            d['x0'] = min(x0 + dx, np.nextafter(x1, -np.inf))
        elif 'r' in edge or edge in ('right',):
            d['x1'] = max(x1 + dx, np.nextafter(x0, np.inf))
        if 't' in edge or edge in ('top',):
            d['y1'] = max(y1 + dy, np.nextafter(y0, np.inf))
        if 'b' in edge or edge in ('bottom',):
            d['y0'] = min(y0 + dy, np.nextafter(y1, -np.inf))

        self._adjust_drag_start = (xd, yd)

        self._rebuild_adjust_ui()
        self._adjust_frame_update()
        self.canvas.draw_idle()

    def _rebuild_adjust_ui(self):
        """刷新 adjust 矩形和手柄的位置（原地更新，不删除重建）"""
        if self._adjust_data is None or self._adjust_rect is None:
            return
        d = self._adjust_data
        self._adjust_rect.set_xy((d['x0'], d['y0']))
        self._adjust_rect.set_width(d['x1'] - d['x0'])
        self._adjust_rect.set_height(d['y1'] - d['y0'])

        # 原地更新 8 个手柄位置，不删除重建
        if not self._adjust_handles:
            self._build_adjust_handles()
            return
        x0, y0, x1, y1 = d['x0'], d['y0'], d['x1'], d['y1']
        xm, ym = (x0 + x1) / 2, (y0 + y1) / 2
        hw, hh = self._handle_size_data()
        new_positions = [
            (x0, y0), (xm, y0), (x1, y0),
            (x0, ym), (x1, ym),
            (x0, y1), (xm, y1), (x1, y1),
        ]
        for rect, (hx, hy) in zip(self._adjust_handles, new_positions):
            rect.set_xy((hx - hw / 2, hy - hh / 2))
            rect.set_width(hw)
            rect.set_height(hh)

    def _detect_adjust_interaction(self, event):
        """返回点击命中的手柄 key，或 'center' / 'outside' / None"""
        if not self._adjust_mode or self._adjust_data is None:
            return None
        if event.xdata is None or event.ydata is None:
            return None

        d = self._adjust_data
        x0, y0, x1, y1 = d['x0'], d['y0'], d['x1'], d['y1']
        xm, ym = (x0 + x1) / 2, (y0 + y1) / 2

        # 在显示坐标中做命中检测（固定像素半径）
        trans = self.ax.transData
        cx, cy = trans.transform((event.xdata, event.ydata))
        r = self._handle_radius

        handles = {
            'bl': trans.transform((x0, y0)), 'bm': trans.transform((xm, y0)),
            'br': trans.transform((x1, y0)), 'ml': trans.transform((x0, ym)),
            'mr': trans.transform((x1, ym)), 'tl': trans.transform((x0, y1)),
            'tm': trans.transform((xm, y1)), 'tr': trans.transform((x1, y1)),
        }

        for key, (hx, hy) in handles.items():
            if abs(cx - hx) <= r and abs(cy - hy) <= r:
                return key

        # 边缘检测（左/右/上/下）
        left_x = handles['bl'][0]
        right_x = handles['br'][0]
        top_y = handles['tl'][1]
        bottom_y = handles['bl'][1]

        if abs(cx - left_x) <= r and bottom_y + r < cy < top_y - r:
            return 'left'
        if abs(cx - right_x) <= r and bottom_y + r < cy < top_y - r:
            return 'right'
        if abs(cy - top_y) <= r and left_x + r < cx < right_x - r:
            return 'top'
        if abs(cy - bottom_y) <= r and left_x + r < cx < right_x - r:
            return 'bottom'

        # 矩形内部
        if x0 < event.xdata < x1 and y0 < event.ydata < y1:
            return 'center'

        return 'outside'

    def _update_adjust_selection_preview(self):
        """根据当前 adjust 窗口实时更新选中预览"""
        if self._adjust_data is None:
            return
        d = self._adjust_data
        self._selection_x_bounds = (d['x0'], d['x1'])
        self._selected_points = self._select_points_in_rect(
            d['x0'], d['y0'], d['x1'], d['y1'])
        self._refresh_visual_style()
        self._update_selection_info()

    def _adjust_frame_update(self):
        """adjust 拖拽每帧专用：只更新高亮图元 + 计数，不碰 16 条曲线的基础属性"""
        if self._adjust_data is None:
            return
        d = self._adjust_data
        self._selection_x_bounds = (d['x0'], d['x1'])
        self._selected_points = self._select_points_in_rect(
            d['x0'], d['y0'], d['x1'], d['y1'])

        # 预计算选中点按曲线分组（一次遍历）
        sel_by_curve = {}
        for c, p in self._selected_points:
            sel_by_curve.setdefault(c, []).append(p)

        effective = self._get_effective_active()

        for idx, curve in enumerate(self.curves):
            is_curve_active = (len(effective) == 0 or idx in effective
                               or idx in self.pinned_curve_indices)
            sel_p_idxs = sel_by_curve.get(idx)

            if is_curve_active and sel_p_idxs:
                xs_p = curve['x'][sel_p_idxs]
                ys_p = curve['y'][sel_p_idxs]
                curve['hl_ctrl_points'].set_data(xs_p, ys_p)
                curve['hl_ctrl_points'].set_alpha(1.0)
                curve['hl_ctrl_points'].set_markersize(7)

                x_fine, y_fine = curve['fine_x_full'], curve['fine_y_full']
                xmin, xmax = self._selection_x_bounds
                mask = (x_fine >= xmin) & (x_fine <= xmax)
                if np.any(mask):
                    curve['hl_spline_line'].set_data(x_fine[mask], y_fine[mask])
                    curve['hl_spline_line'].set_alpha(1.0)
                    curve['hl_spline_line'].set_linewidth(4.0)
                else:
                    curve['hl_spline_line'].set_data([], [])
            else:
                curve['hl_spline_line'].set_data([], [])
                curve['hl_ctrl_points'].set_data([], [])

        # 轻量更新底部计数（不调 _update_selection_info 避免重复遍历 + _draw_idle）
        n = len(self._selected_points)
        self.info_num.set_text(f"{n} pts")
        if sel_by_curve:
            detail = ", ".join(f"C{c}: {len(p_)}p" for c, p_ in sorted(sel_by_curve.items()))
            self.info_tag.set_text(f"Adjusting  ({detail})")
        else:
            self.info_tag.set_text("Adjusting...")

    def _delete_selected_points(self):
        if not self._selected_points:
            return
        by_curve = {}
        for c, p in self._selected_points:
            by_curve.setdefault(c, []).append(p)

        deleted_any = False
        for c_idx, p_idxs in by_curve.items():
            curve = self.curves[c_idx]
            keep = [i for i in range(len(curve['x'])) if i not in p_idxs]
            if len(keep) < 4:
                print(f"  {curve['name']} 删除后不足 4 点，跳过")
                continue
            curve['x'] = np.arange(len(keep), dtype=float)
            curve['y'] = curve['y'][keep]
            self._update_spline(c_idx)
            deleted_any = True

        if deleted_any:
            self._save_current_state()
        self._clear_selection()

    def _do_resample(self, event=None):
        """核心业务功能：将选中曲线的局部控制点，平滑增点/降点重采样为 Y 个点"""
        if not hasattr(self, '_resample_box') or not self._resample_box:
            return

        try:
            n_new = int(self._resample_box.text)
        except ValueError:
            messagebox.showwarning("输入错误", "请输入有效的段内目标点数整数（如 10）")
            return
        if n_new < 4:
            messagebox.showwarning("约束错误", "局部段重采样后至少需要保留 4 个控制点以维持三次样条拟合")
            return

        effective = self._get_effective_active()
        by_curve = {}
        if self._selected_points:
            for c, p in self._selected_points:
                if len(effective) == 0 or c in effective:
                    by_curve.setdefault(c, []).append(p)
        else:
            targets = sorted(effective) if len(effective) > 0 else list(range(len(self.curves)))
            for c in targets:
                by_curve[c] = list(range(len(self.curves[c]['y'])))

        if not by_curve:
            print("当前未选中任何有效曲线或控制点")
            return

        counts = {len(p_idxs) for p_idxs in by_curve.values()}
        if len(counts) > 1:
            messagebox.showerror("重采样失败", f"强约束未满足！当前框选的各曲线内部点数不一致: {counts}，无法批量重采样。")
            return

        modified_any = False

        for c_idx, p_idxs in by_curve.items():
            curve = self.curves[c_idx]
            orig_y = curve['y']
            total_pts = len(orig_y)

            i_min, i_max = min(p_idxs), max(p_idxs)
            n_old = i_max - i_min + 1

            if n_old == n_new and self._selected_points:
                continue

            pad = 2
            ext_min = max(0, i_min - pad)
            ext_max = min(total_pts - 1, i_max + pad)

            seg_y = orig_y[ext_min:ext_max + 1]
            seg_x = np.arange(len(seg_y), dtype=float)

            k_order = 3 if len(seg_y) > 3 else (len(seg_y) - 1)
            if k_order < 1:
                continue

            tck, u = splprep([seg_x, seg_y], s=0, k=k_order)

            left_pad_count = i_min - ext_min
            right_pad_count = ext_max - i_max
            total_seg_new_count = left_pad_count + n_new + right_pad_count

            u_new = np.linspace(0, 1, total_seg_new_count)
            _, y_resampled_ext = splev(u_new, tck)

            start_idx = left_pad_count
            end_idx = total_seg_new_count - right_pad_count
            y_resampled_core = y_resampled_ext[start_idx:end_idx]

            y_new = np.concatenate([
                orig_y[:i_min],
                y_resampled_core,
                orig_y[i_max + 1:]
            ])

            curve['x'] = np.arange(len(y_new), dtype=float)
            curve['y'] = y_new
            self._update_spline(c_idx)
            modified_any = True

        if modified_any:
            self._save_current_state()
            print(f"成功！已将各曲线选中段的控制点平滑转换为了 {n_new} 个点。")

        self._clear_selection()

    # ===================== C / V 复制粘贴 =====================

    def _copy_selected(self):
        """C: 复制选中控制点的 Y 值（限单条曲线）"""
        if not self._selected_points:
            return

        # 必须所有选中点属于同一条曲线
        curves = {c for c, _ in self._selected_points}
        if len(curves) != 1:
            print("  ⚠ 复制仅支持单条曲线，当前选中跨多条曲线")
            return

        c_idx = curves.pop()
        if c_idx in self.pinned_curve_indices:
            return

        # 按点索引升序提取 Y 值
        sorted_pts = sorted(p for c, p in self._selected_points if c == c_idx)
        y_vals = [self.curves[c_idx]['y'][p] for p in sorted_pts]

        self._clipboard = {'curve_idx': c_idx, 'y_values': y_vals}
        print(f"  📋 已复制 {self.curves[c_idx]['name']} 的 {len(y_vals)} 个点")

    def _paste_to_curve(self):
        """Ctrl+V: 粘贴到当前活跃曲线（或原曲线）最后选中点之后"""
        if self._clipboard is None:
            return

        y_vals = self._clipboard['y_values']

        # 目标曲线：当前活跃的单条曲线优先；否则用复制来源曲线
        effective = self._get_effective_active()
        if len(effective) == 1:
            c_idx = next(iter(effective))
        else:
            c_idx = self._clipboard['curve_idx']

        if c_idx >= len(self.curves) or c_idx in self.pinned_curve_indices:
            return

        curve = self.curves[c_idx]

        # 确定插入位置：当前选中点中属于目标曲线的最大索引 + 1；否则末尾
        same_curve_pts = [p for c, p in self._selected_points if c == c_idx]
        if same_curve_pts:
            insert_at = max(same_curve_pts) + 1
        else:
            insert_at = len(curve['y'])

        # 插入
        new_y = np.insert(curve['y'], insert_at, y_vals)
        curve['y'] = new_y
        curve['x'] = np.arange(len(new_y), dtype=float)
        self._update_spline(c_idx)
        self._save_current_state()

        # 更新选中：清除旧选区，选中新插入的点
        new_selected = [(c_idx, insert_at + i) for i in range(len(y_vals))]
        self._selected_points = new_selected
        self._refresh_visual_style()
        self._update_selection_info()

        from_curve = self.curves[self._clipboard['curve_idx']]['name']
        if c_idx != self._clipboard['curve_idx']:
            print(f"  📋 已粘贴 {len(y_vals)} 个点: {from_curve} → {curve['name']} [{insert_at}]")
        else:
            print(f"  📋 已粘贴 {len(y_vals)} 个点到 {curve['name']} [{insert_at}]")
        """核心业务功能：将选中曲线的局部控制点，平滑增点/降点重采样为 Y 个点"""
        if not hasattr(self, '_resample_box') or not self._resample_box:
            return
        
        try:
            n_new = int(self._resample_box.text)
        except ValueError:
            messagebox.showwarning("输入错误", "请输入有效的段内目标点数整数（如 10）")
            return
        if n_new < 4:
            messagebox.showwarning("约束错误", "局部段重采样后至少需要保留 4 个控制点以维持三次样条拟合")
            return

        # 1. 确定当前哪些曲线和哪些点将被作为目标处理
        effective = self._get_effective_active()
        by_curve = {}
        if self._selected_points:
            for c, p in self._selected_points:
                if len(effective) == 0 or c in effective:
                    by_curve.setdefault(c, []).append(p)
        else:
            targets = sorted(effective) if len(effective) > 0 else list(range(len(self.curves)))
            for c in targets:
                by_curve[c] = list(range(len(self.curves[c]['y'])))

        if not by_curve:
            print("当前未选中任何有效曲线或控制点")
            return

        # 2. 【强约束判定】确保当前所有选中的曲线，它们被选中的点数（X）必须完全一致
        counts = {len(p_idxs) for p_idxs in by_curve.values()}
        if len(counts) > 1:
            messagebox.showerror("重采样失败", f"强约束未满足！当前框选的各曲线内部点数不一致: {counts}，无法批量重采样。")
            return

        modified_any = False

        # 3. 开始执行带边界保护的三次样条局部重采样
        for c_idx, p_idxs in by_curve.items():
            curve = self.curves[c_idx]
            orig_y = curve['y']
            total_pts = len(orig_y)

            i_min, i_max = min(p_idxs), max(p_idxs)
            n_old = i_max - i_min + 1

            if n_old == n_new and self._selected_points:
                continue  # 数量一样且是局部框选，不作变动

            # 为保证交接处极度平滑，向左右未选中区外扩 2 个控制点作为“缓冲约束区”
            pad = 2
            ext_min = max(0, i_min - pad)
            ext_max = min(total_pts - 1, i_max + pad)

            seg_y = orig_y[ext_min:ext_max + 1]
            seg_x = np.arange(len(seg_y), dtype=float)

            k_order = 3 if len(seg_y) > 3 else (len(seg_y) - 1)
            if k_order < 1:
                continue

            # 构建样条核心参数
            tck, u = splprep([seg_x, seg_y], s=0, k=k_order)

            # 计算生成的新段里包含的缓冲区节点数
            left_pad_count = i_min - ext_min
            right_pad_count = ext_max - i_max
            total_seg_new_count = left_pad_count + n_new + right_pad_count

            u_new = np.linspace(0, 1, total_seg_new_count)
            _, y_resampled_ext = splev(u_new, tck)

            # 剔除缓冲区，完美截取并恢复我们需要的 Y 个新点
            start_idx = left_pad_count
            end_idx = total_seg_new_count - right_pad_count
            y_resampled_core = y_resampled_ext[start_idx:end_idx]

            # 完美拼回大数组
            y_new = np.concatenate([
                orig_y[:i_min],
                y_resampled_core,
                orig_y[i_max + 1:]
            ])

            # 重新构建自增一维 X 轴索引并更新艺术家图元
            curve['x'] = np.arange(len(y_new), dtype=float)
            curve['y'] = y_new
            self._update_spline(c_idx)
            modified_any = True

        if modified_any:
            self._save_current_state()
            print(f"成功！已将各曲线选中段的控制点平滑转换为了 {n_new} 个点。")
        
        # 联动重置清除框选状态
        self._clear_selection()

    def add_curve(self, x_ctrl, y_ctrl, color='blue', name="Curve"):
        x = np.array(x_ctrl, dtype=float)
        y = np.array(y_ctrl, dtype=float)
        if len(x) < 4:
            raise ValueError("At least 4 control points required.")
        ctrl_points, = self.ax.plot(x, y, 'o', color=color, markersize=5, alpha=0.65, zorder=2, markeredgewidth=0)
        spline_line, = self.ax.plot([], [], '-', color=color, linewidth=2.8, label=name, zorder=1)
        hl_ctrl_points, = self.ax.plot([], [], 'o', color=color, markersize=8, alpha=0.0, zorder=4, markeredgewidth=0.8, markeredgecolor='#222')
        hl_spline_line, = self.ax.plot([], [], '-', color=color, linewidth=5.0, alpha=0.0, zorder=3)
        curve_dict = {
            'x': x, 'y': y,
            'color': color,
            'ctrl_points': ctrl_points,
            'spline_line': spline_line,
            'hl_ctrl_points': hl_ctrl_points,
            'hl_spline_line': hl_spline_line,
            'fine_x_full': None, 'fine_y_full': None,
            'name': name
        }
        self.curves.append(curve_dict)
        self._update_spline(len(self.curves) - 1)

    # ===================== 文件 I/O =====================

    def _clear_all_curves(self):
        for curve in self.curves:
            for key in ['ctrl_points', 'spline_line', 'hl_ctrl_points', 'hl_spline_line']:
                if curve.get(key) is not None:
                    curve[key].remove()
        self.curves.clear()
        self.undo_stack.clear()
        self._selected_points.clear()
        self._batch_moving = False
        self._selection_x_bounds = None
        self._slider_base_state = None
        self.active_curve_indices = set()
        self.locked_curve_indices = set()
        self.pinned_curve_indices = set()
        self.current_label = "All"
        self._active_curve_idx = None
        self._active_point_idx = None
        self._tab_index = 0
        self._last_npts_val = None
        self._hide_drop_hint()
        self.ax.relim()
        self.ax.autoscale_view()

    def _auto_range_axes(self, event=None):
        if not self.curves:
            return
        all_x = np.concatenate([c['x'] for c in self.curves])
        all_y = np.concatenate([c['y'] for c in self.curves])
        if len(all_x) == 0 or len(all_y) == 0:
            return

        x_min, x_max = np.min(all_x), np.max(all_x)
        y_min, y_max = np.min(all_y), np.max(all_y)

        x_range = x_max - x_min or 1.0
        y_range = y_max - y_min or 1.0
        margin = 0.08

        self.ax.set_xlim(x_min - x_range * margin, x_max + x_range * margin)
        self.ax.set_ylim(y_min - y_range * margin, y_max + y_range * margin)
        self.canvas.draw_idle()

    def _expand_axes_if_needed(self):
        """仅在拖拽点超出当前 Y 轴范围时才扩，不超出则完全不改视图"""
        if not self.curves:
            return
        all_y = np.concatenate([c['y'] for c in self.curves])
        if len(all_y) == 0:
            return

        y_min, y_max = np.min(all_y), np.max(all_y)
        cur_ylo, cur_yhi = self.ax.get_ylim()

        new_lo, new_hi = cur_ylo, cur_yhi
        if y_min < cur_ylo:
            new_lo = y_min
        if y_max > cur_yhi:
            new_hi = y_max

        if new_lo != cur_ylo or new_hi != cur_yhi:
            self.ax.set_ylim(new_lo, new_hi)
            self.canvas.draw_idle()

    def _detect_format(self, file_path):
        ext = splitext(file_path)[1].lower()
        if ext == '.csv':
            return 'csv'
        elif ext in ('.xlsx', '.xls'):
            return 'excel'
        elif ext == '.npy':
            return 'npy'
        else:
            raise ValueError(f"不支持的文件格式: {ext}（支持 .csv .xlsx .npy）")

    def _load_data_from_file(self, file_path):
        """加载文件，返回 (curves_list, is_ragged)。
        curves_list: [[y1,y2,...], ...] 每条曲线一个列表
        is_ragged: True 表示曲线长度不一致"""
        fmt = self._detect_format(file_path)

        if fmt == 'npy':
            data = np.load(file_path, allow_pickle=True)
            if data.ndim == 1:
                data = data.reshape(-1, 1)
            if data.ndim != 2:
                raise ValueError(f"NPY 数据必须是 1D 或 2D，当前 shape: {data.shape}")
            n_rows, n_cols = data.shape
            if n_rows < n_cols:
                data = data.T
            if data.shape[0] < 4:
                raise ValueError(f"每条曲线至少需要 4 个采样点，当前仅 {data.shape[0]} 行")
            return [data[:, i].tolist() for i in range(data.shape[1])], False

        elif fmt == 'csv':
            # 逐行读取，支持不等长列
            curves_by_col = []
            with open(file_path, 'r') as f:
                for line in f:
                    line = line.strip()
                    if not line:
                        continue
                    vals = [float(x) for x in line.split(',') if x.strip()]
                    for ci, v in enumerate(vals):
                        while len(curves_by_col) <= ci:
                            curves_by_col.append([])
                        curves_by_col[ci].append(v)
            if not curves_by_col:
                raise ValueError("CSV 文件为空")
            # 方向检测：列数 > 每列行数 → 可能需要转置
            if len(curves_by_col) > max(len(c) for c in curves_by_col):
                # 每行作为一条曲线
                curves_by_col = [[r] for r in curves_by_col[0]]
                # Re-read: each row is a curve
                with open(file_path, 'r') as f:
                    curves_by_col = []
                    for line in f:
                        line = line.strip()
                        if not line:
                            continue
                        vals = [float(x) for x in line.split(',') if x.strip()]
                        if vals:
                            curves_by_col.append(vals)
            for c in curves_by_col:
                if len(c) < 4:
                    raise ValueError(f"每条曲线至少需要 4 个采样点，当前某曲线仅 {len(c)} 点")
            is_ragged = len({len(c) for c in curves_by_col}) > 1
            return curves_by_col, is_ragged

        elif fmt == 'excel':
            if not HAS_EXCEL:
                raise ImportError("读取 Excel 需要 openpyxl: pip install openpyxl")
            wb = load_workbook(file_path, data_only=True)
            ws = wb.active
            rows = [[cell.value for cell in row] for row in ws.iter_rows()]
            wb.close()
            if not rows:
                raise ValueError("Excel 文件为空")

            # 逐列读取，只取非空值，不补 0
            max_cols = max(len(r) for r in rows) if rows else 0
            curves_by_col = []
            for ci in range(max_cols):
                col = [float(r[ci]) for r in rows
                       if ci < len(r) and r[ci] is not None]
                if col:
                    curves_by_col.append(col)

            # 方向检测：列少行多 → 每列是一条曲线；列多行少 → 转置
            if len(curves_by_col) > sum(len(c) for c in curves_by_col) / len(curves_by_col) * 2:
                curves_by_col = [[float(v) for v in r if v is not None]
                                 for r in rows if any(v is not None for v in r)]

            for c in curves_by_col:
                if len(c) < 4:
                    raise ValueError(f"每条曲线至少需要 4 个采样点，当前某曲线仅 {len(c)} 点")
            is_ragged = len({len(c) for c in curves_by_col}) > 1
            return curves_by_col, is_ragged

        raise ValueError(f"不支持的文件格式: {fmt}")

    def load_curves_from_file(self, file_path=None):
        if file_path is None:
            file_path = filedialog.askopenfilename(
                title="打开曲线文件",
                filetypes=[
                    ("所有支持格式", "*.csv;*.xlsx;*.xls;*.npy"),
                    ("CSV 文件", "*.csv"),
                    ("Excel 文件", "*.xlsx;*.xls"),
                    ("NumPy 文件", "*.npy"),
                ]
            )
            if not file_path:
                return

        try:
            curves_list, is_ragged = self._load_data_from_file(file_path)
        except Exception as e:
            messagebox.showerror("加载错误", f"无法加载文件:\n{e}")
            return

        self.current_file_path = file_path
        self._clear_all_curves()
        self._clear_radio_panel()

        curve_colors = []
        for idx, y_data in enumerate(curves_list):
            color = self.color_list[idx % len(self.color_list)] if self.color_list else f"C{idx}"
            name = f"Curve {idx}"
            curve_colors.append(color)
            y_arr = np.array(y_data, dtype=float)
            self.add_curve(np.arange(len(y_arr)), y_arr, color=color, name=name)

        self._build_radio_panel(curve_colors)
        self._auto_range_axes()
        self._save_current_state()
        self._reset_slider_widgets()
        self._hide_drop_hint()
        self._update_selection_info()

        fname = basename(file_path)
        self.fig.canvas.manager.set_window_title(f"CurveMagician - {fname}")
        n_curves = len(curves_list)
        if is_ragged:
            pts_info = ", ".join(f"{len(c)}" for c in curves_list)
            print(f"已加载 {n_curves} 条曲线 (不等长: {pts_info})，来自 {file_path}")
        else:
            print(f"已加载 {n_curves} 条曲线 ({len(curves_list[0])} 点/条)，来自 {file_path}")

    def save_file(self, event=None):
        if not self.curves:
            messagebox.showwarning("保存", "没有曲线数据可保存。")
            return

        # 检测曲线长度是否一致
        lengths = [len(c['y']) for c in self.curves]
        is_ragged = len(set(lengths)) > 1

        if is_ragged:
            default_ext = ".xlsx"
            initial_file = "curves.xlsx"
            filetypes = [
                ("Excel 文件", "*.xlsx"),
                ("CSV 文件", "*.csv"),
            ]
        else:
            default_ext = ".npy"
            initial_file = "curves.npy"
            filetypes = [
                ("NumPy 文件", "*.npy"),
                ("CSV 文件", "*.csv"),
                ("Excel 文件", "*.xlsx"),
            ]

        if self.current_file_path:
            default_ext = splitext(self.current_file_path)[1]
            initial_file = basename(self.current_file_path)

        file_path = filedialog.asksaveasfilename(
            title="保存曲线为",
            initialfile=initial_file,
            defaultextension=default_ext,
            filetypes=filetypes,
        )
        if not file_path:
            return

        try:
            fmt = self._detect_format(file_path)

            if is_ragged and fmt == 'npy':
                messagebox.showerror("保存错误",
                    "曲线长度不一致，无法保存为 NPY。请选择 Excel 或 CSV 格式。")
                return

            if fmt == 'npy':
                data = np.array([curve['y'] for curve in self.curves]).T
                np.save(file_path, data)
            elif fmt == 'csv':
                max_len = max(lengths)
                with open(file_path, 'w') as f:
                    for row_idx in range(max_len):
                        row_vals = []
                        for c in self.curves:
                            if row_idx < len(c['y']):
                                row_vals.append(f"{c['y'][row_idx]:.8g}")
                            else:
                                row_vals.append("")
                        f.write(",".join(row_vals) + "\n")
            elif fmt == 'excel':
                if not HAS_EXCEL:
                    raise ImportError("保存 Excel 需要 openpyxl: pip install openpyxl")
                wb = Workbook()
                ws = wb.active
                max_len = max(lengths)
                for row_idx in range(max_len):
                    row = []
                    for c in self.curves:
                        if row_idx < len(c['y']):
                            row.append(c['y'][row_idx])
                        else:
                            row.append(None)
                    ws.append(row)
                wb.save(file_path)

            self.current_file_path = file_path
            fname = basename(file_path)
            self.fig.canvas.manager.set_window_title(f"CurveMagician - {fname}")
            if is_ragged:
                print(f"已保存 {len(self.curves)} 条曲线 (不等长) 到 {file_path}")
            else:
                print(f"已保存 {len(self.curves)} 条曲线 ({lengths[0]} 点/条) 到 {file_path}")
        except Exception as e:
            messagebox.showerror("保存错误", f"保存失败:\n{e}")

    def on_open_clicked(self, event=None):
        self.load_curves_from_file()

    # ===================== 单选面板构建 =====================

    def _clear_radio_panel(self):
        if self.ax_radio is None:
            return
        self.ax_radio.cla()
        self.ax_radio.set_facecolor('#f2f4f7')
        self.ax_radio.set_xticks([])
        self.ax_radio.set_yticks([])
        self.ax_radio.set_navigate(False)
        self.ax_radio.set_xlim(0, 1)
        self.ax_radio.set_ylim(0, 1)
        self._top_dots = None
        self._radio_circles.clear()
        self._color_patches.clear()
        self._lock_icons.clear()
        self.radio_menu = None
        self.radio_labels = []

    def _build_radio_panel(self, curve_colors):
        """构建多选 checkbox 面板：All + 各曲线，点击 toggle 勾选"""
        if self.ax_radio is None:
            return

        ax = self.ax_radio
        ax.set_facecolor('#f2f4f7')
        base_labels = ["All"] + [self.curves[i]['name'] for i in range(len(curve_colors))]
        num_labels = len(base_labels)
        font_size = 9 if num_labels > 8 else 10

        self._radio_ys = np.linspace(1, 0, num_labels + 2)[1:-1]
        self._radio_labels = base_labels
        self._radio_colors = ['#2E86DE'] + list(curve_colors)  # All 用亮蓝
        dot_radius = 0.028 if num_labels > 8 else 0.035
        patch_h = 0.55 / num_labels if num_labels > 8 else 0.038

        self._radio_circles = []
        self._lock_icons = []
        self._color_patches.clear()

        # 断开旧的 radio click handler，避免重复注册互相抵消
        if hasattr(self, '_radio_click_cid') and self._radio_click_cid is not None:
            self.fig.canvas.mpl_disconnect(self._radio_click_cid)
            self._radio_click_cid = None

        for i in range(num_labels):
            y = self._radio_ys[i]

            # 方框 — checkbox 风格
            c = Rectangle(
                (0.06, y - dot_radius), dot_radius * 2, dot_radius * 2,
                transform=ax.transAxes,
                facecolor='white',
                edgecolor='#9098a8',
                linewidth=2.0,
                zorder=10,
            )
            ax.add_patch(c)
            self._radio_circles.append(c)

            # 锁图标 — checkbox 中心，初始隐藏，用 emoji 字体
            icon = ax.text(
                0.06 + dot_radius, y, '',
                transform=ax.transAxes,
                fontsize=font_size + 10, fontfamily='Segoe UI Emoji',
                ha='center', va='center',
                zorder=15, visible=False,
            )
            self._lock_icons.append(icon)

            # 曲线名称
            ax.text(
                0.42, y, base_labels[i],
                transform=ax.transAxes,
                fontsize=font_size,
                fontweight='bold' if i == 0 else 'normal',
                color='#2d3436',
                va='center',
            )

            if i == 0:
                continue

            # 色块
            rect = Rectangle(
                (0.18, y - patch_h / 2), 0.20, patch_h,
                facecolor=curve_colors[i - 1],
                transform=ax.transAxes,
                zorder=0,
                edgecolor='#c0c7d0',
                linewidth=1.2,
                alpha=1.0,
            )
            ax.add_patch(rect)
            self._color_patches.append(rect)

        self.radio_labels = base_labels
        self._refresh_radio_panel()

        def _on_radio_click(event):
            if event.inaxes != ax or event.button != 1:
                return
            if event.ydata is None or not hasattr(self, '_radio_ys'):
                return
            dists = np.abs(self._radio_ys - event.ydata)
            closest = int(np.argmin(dists))
            if dists[closest] < 0.8 / num_labels:
                if event.dblclick:
                    self.toggle_active_curve(self._radio_labels[closest])
                else:
                    self.set_active_curve(self._radio_labels[closest])

        self._radio_click_cid = self.fig.canvas.mpl_connect(
            'button_press_event', _on_radio_click)

    def _refresh_radio_panel(self):
        """更新 checkbox 状态：实心=单击选中，边框加粗+对勾=双击锁定"""
        if not hasattr(self, '_radio_circles') or not self._radio_circles:
            return
        effective = self._get_effective_active()
        is_all = len(effective) == 0
        for i, c in enumerate(self._radio_circles):
            if i == 0:
                checked = is_all
                locked = False
                pinned = False
            else:
                checked = (i - 1) in self.active_curve_indices
                locked = (i - 1) in self.locked_curve_indices
                pinned = (i - 1) in self.pinned_curve_indices
            if pinned or locked or checked:
                c.set_facecolor(self._radio_colors[i])
            else:
                c.set_facecolor('white')
            c.set_edgecolor('#9098a8')
            c.set_linewidth(2.0)

        # 更新图标 — 🔒 = 锁定  📌 = 固定
        for i, icon in enumerate(self._lock_icons):
            if i == 0:
                icon.set_visible(False)
            elif (i - 1) in self.pinned_curve_indices:
                icon.set_text('\U0001F4CC'); icon.set_visible(True)
            elif (i - 1) in self.locked_curve_indices:
                icon.set_text('\U0001F512'); icon.set_visible(True)
            elif (i - 1) in self.active_curve_indices:
                icon.set_visible(False)
            else:
                icon.set_visible(False)
        self.fig.canvas.draw_idle()

    # ===================== 拖拽支持 =====================

    def _show_drop_hint(self):
        if self.drop_hint is None:
            self.drop_hint = self.ax.text(
                0.5, 0.5,
                '拖放 CSV / Excel / NPY 文件到此处\n或点击 Open 按钮选择文件',
                transform=self.ax.transAxes,
                ha='center', va='center',
                fontsize=15, color='#636e72', alpha=0.35,
                fontweight='bold',
                zorder=100,
            )
        else:
            self.drop_hint.set_visible(True)
        self.canvas.draw_idle()

    def _hide_drop_hint(self):
        if self.drop_hint is not None:
            self.drop_hint.set_visible(False)
            self.canvas.draw_idle()

    def _on_drop(self, event):
        files = self.fig.canvas.manager.window.tk.splitlist(event.data)
        if not files:
            return

        cleaned = []
        for f in files:
            f = f.strip()
            if f.startswith('{') and f.endswith('}'):
                f = f[1:-1]
            cleaned.append(f)

        for fpath in cleaned:
            try:
                self._detect_format(fpath)
                self.load_curves_from_file(fpath)
                return
            except ValueError:
                continue

        messagebox.showwarning(
            "不支持的文件",
            "未找到支持的文件格式。\n支持: .csv, .xlsx, .npy"
        )

    def setup_drag_and_drop(self):
        if not HAS_DND:
            return
        try:
            import tkinter as tk
            root = tk._default_root or self.fig.canvas.manager.window.winfo_toplevel()
            TkinterDnD.require(root)
            canvas_widget = self.fig.canvas.get_tk_widget()
            canvas_widget.drop_target_register(DND_FILES)
            canvas_widget.dnd_bind('<<Drop>>', self._on_drop)
            print("文件拖拽功能已启用")
        except Exception as e:
            print(f"拖拽注册失败: {e}")

    def _update_spline(self, curve_idx):
        curve = self.curves[curve_idx]
        x, y = curve['x'], curve['y']
        curve['ctrl_points'].set_data(x, y)
        curve['spline_line'].set_data(x, y)
        curve['fine_x_full'] = x
        curve['fine_y_full'] = y

    def on_press(self, event):
        if self._is_toolbar_active():
            return

        if not hasattr(event, 'inaxes') or event.inaxes is None:
            return

        slider_axes = []
        if self.slider_scale: slider_axes.append(self.slider_scale.ax)
        if self.slider_smooth: slider_axes.append(self.slider_smooth.ax)
        if self.slider_noise: slider_axes.append(self.slider_noise.ax)

        if event.inaxes in slider_axes:
            self._is_sliding = True
            self._noise_seed = np.random.randint(0, 2**31)
            return

        if event.inaxes != self.ax:
            return

        # ── 可调窗口模式下的交互 ──
        if self._adjust_mode and event.button == 1:
            # 双击矩形内部 → 确认
            if getattr(event, 'dblclick', False):
                edge = self._detect_adjust_interaction(event)
                if edge in ('center', 'left', 'right', 'top', 'bottom',
                            'tl', 'tr', 'bl', 'br', 'tm', 'bm', 'ml', 'mr'):
                    self._confirm_adjust()
                    return
            edge = self._detect_adjust_interaction(event)
            if edge == 'outside':
                self._cancel_adjust()
                return
            if edge is not None:
                self._adjust_drag_edge = edge
                self._adjust_drag_start = (event.xdata, event.ydata)
                return
            return

        if event.button == 3:
            if self._adjust_mode:
                self._cancel_adjust()
                return
            self._clear_selection()
            return
        if event.button == 1:
            if self._selected_points:
                self._batch_moving = True
                self._batch_origin = (event.xdata, event.ydata)
                self._in_continuous_drag = True
                return
            c_idx, p_idx = self._get_closest_point(event)
            if c_idx is not None and p_idx is not None:
                self._active_curve_idx = c_idx
                self._active_point_idx = p_idx
                self._in_continuous_drag = True
                self.curves[c_idx]['ctrl_points'].set_markersize(8)
                self.curves[c_idx]['ctrl_points'].set_alpha(1.0)
                self.canvas.draw_idle()
                return
            # 普通框选 / Shift 框选
            self._selecting = True
            self._select_start = (event.xdata, event.ydata)
            self._shift_selecting = (hasattr(event, 'key') and event.key == 'shift')
            self._select_rect = Rectangle(self._select_start, 0, 0, fill=True, alpha=0.2, color='gray')
            self.ax.add_patch(self._select_rect)

    def on_motion(self, event):
        if self._is_toolbar_active():
            return
        if self._is_sliding:
            return

        # ── 可调窗口拖拽：必须在 axes 检查之前，因为上下拖拽容易出界 ──
        if self._adjust_mode and self._adjust_drag_edge is not None:
            self._handle_adjust_drag(event)
            return

        if not hasattr(event, 'inaxes') or event.inaxes != self.ax or event.ydata is None:
            return

        if self._active_curve_idx is not None:
            c_idx = self._active_curve_idx
            p_center = self._active_point_idx
            curve = self.curves[c_idx]
            pts_idx = self._get_linked_point_indices(len(curve['y']), p_center)
            dy = event.ydata - curve['y'][p_center]
            for pid in pts_idx:
                curve['y'][pid] += dy
            self._update_spline(c_idx)
            if self._selection_x_bounds is not None:
                self._refresh_visual_style()
            else:
                self.canvas.draw_idle()
            return
        if self._selecting and self._select_rect is not None:
            x0, y0 = self._select_start
            x1, y1 = event.xdata, event.ydata
            self._select_rect.set_width(x1 - x0)
            self._select_rect.set_height(y1 - y0)
            self.canvas.draw_idle()
            return
        if self._batch_moving and self._batch_origin is not None:
            dy = event.ydata - self._batch_origin[1]
            effective = self._get_effective_active()
            for c_idx, p_idx in self._selected_points:
                if len(effective) == 0 or c_idx in effective:
                    self.curves[c_idx]['y'][p_idx] += dy
            affected = set(c for c, _ in self._selected_points)
            for c in affected:
                self._update_spline(c)
            self._batch_origin = (self._batch_origin[0], event.ydata)
            self._refresh_visual_style()

    def on_release(self, event):
        # ── 可调窗口拖拽手柄松手 — 停止拖拽但保持 adjust 模式 ──
        if self._adjust_mode and self._adjust_drag_edge is not None:
            self._adjust_drag_edge = None
            self._adjust_drag_start = None
            return

        if self._selecting:
            self._selecting = False
            was_shift = self._shift_selecting
            self._shift_selecting = False
            if (self._select_start and event.xdata is not None and event.ydata is not None):
                x0, y0 = self._select_start
                x1, y1 = event.xdata, event.ydata
                if was_shift:
                    # Shift+框选 → 进入可调窗口模式，先清理拖拽矩形
                    if self._select_rect is not None:
                        self._select_rect.remove()
                        self._select_rect = None
                    self._enter_adjust_mode(x0, y0, x1, y1)
                    return
                else:
                    self._selection_x_bounds = (min(x0, x1), max(x0, x1))
                    self._selected_points = self._select_points_in_rect(x0, y0, x1, y1)
                    self._refresh_visual_style()
            if self._select_rect is not None:
                self._select_rect.remove()
                self._select_rect = None
            self._update_selection_info()

        if self._is_toolbar_active():
            return

        if self._is_sliding:
            self._is_sliding = False
            self._noise_seed = None
            self._save_current_state()
            self._slider_base_state = [{'y': crv['y'].copy()} for crv in self.curves]
            return

        if not hasattr(event, 'inaxes') or event.inaxes is None:
            return

        need_save = False
        if self._active_curve_idx is not None:
            need_save = True
            self._refresh_visual_style()

        if self._batch_moving:
            need_save = True

        self._active_curve_idx = None
        self._active_point_idx = None
        self._batch_moving = False
        self._batch_origin = None
        self._in_continuous_drag = False

        if need_save:
            self._save_current_state()
            self._reset_slider_widgets()
            self._expand_axes_if_needed()

        self._update_selection_info()

    def _get_closest_point(self, event):
        if not self.curves or event.xdata is None or event.ydata is None:
            return None, None
        mouse_xy = self.ax.transData.transform((event.xdata, event.ydata))
        effective = self._get_effective_active()
        if len(effective) > 0:
            candidates = sorted(effective)
        else:
            candidates = list(range(len(self.curves)))
        # 排除 pinned 曲线 — 永不被选中/拖拽
        candidates = [c for c in candidates if c not in self.pinned_curve_indices]
        if not candidates:
            return None, None
        min_dist = float('inf')
        closest_curve, closest_point = None, None
        for c_idx in candidates:
            curve = self.curves[c_idx]
            x, y = curve['x'], curve['y']
            pts_xy = self.ax.transData.transform(np.c_[x, y])
            distances = np.linalg.norm(pts_xy - mouse_xy, axis=1)
            p_idx = np.argmin(distances)
            if distances[p_idx] < min_dist:
                min_dist = distances[p_idx]
                closest_curve, closest_point = c_idx, p_idx
        if min_dist < self._epsilon:
            return closest_curve, closest_point
        return None, None

    def _select_points_in_rect(self, x0, y0, x1, y1):
        selected = []
        xmin, xmax = min(x0, x1), max(x0, x1)
        ymin, ymax = min(y0, y1), max(y0, y1)
        effective = self._get_effective_active()
        if len(effective) > 0:
            curves_to_check = [(i, self.curves[i]) for i in sorted(effective)]
        else:
            curves_to_check = list(enumerate(self.curves))
        # 排除 pinned 曲线 — 永不被框选
        curves_to_check = [(c, crv) for c, crv in curves_to_check
                           if c not in self.pinned_curve_indices]
        for c_idx, curve in curves_to_check:
            xs, ys = curve['x'], curve['y']
            # numpy 矢量化 — 对 300 点一枪头判定，避免 Python 逐点循环
            mask = (xs >= xmin) & (xs <= xmax) & (ys >= ymin) & (ys <= ymax)
            for p_idx in np.flatnonzero(mask):
                selected.append((c_idx, int(p_idx)))
        return selected

    def _get_linked_point_indices(self, total_len, center_idx):
        n = self.neighbor_num
        return list(range(max(0, center_idx - n), min(total_len - 1, center_idx + n) + 1))

    def set_neighbor_num(self, text):
        try:
            val = int(text)
            self.neighbor_num = max(0, val)
            self._update_selection_info()
        except ValueError:
            print("Please enter an integer!")

    def _save_current_state(self):
        state = [{'x': crv['x'].copy(), 'y': crv['y'].copy()} for crv in self.curves]
        self.undo_stack.append(state)

    def undo(self, event=None):
        if len(self.undo_stack) <= 1:
            return
        self.undo_stack.pop()
        prev_state = self.undo_stack[-1]
        for idx, crv in enumerate(self.curves):
            if idx < len(prev_state):
                crv['x'] = prev_state[idx]['x'].copy()
                crv['y'] = prev_state[idx]['y'].copy()
                self._update_spline(idx)
        self._refresh_visual_style()
        self._reset_slider_widgets()

    def on_key(self, event):
        # ── 可调窗口模式下的键盘操作 ──
        if self._adjust_mode:
            if event.key == 'escape':
                self._cancel_adjust()
                return
            if event.key == 'enter':
                self._confirm_adjust()
                return
            # adjust 模式下屏蔽其他按键
            return

        if event.key in ('c', 'ctrl+c'):
            self._copy_selected()
        elif event.key in ('v', 'ctrl+v'):
            self._paste_to_curve()
        elif event.key == 's':
            self.save_file()
        elif event.key == 'o':
            self.on_open_clicked()
        elif event.key == 'r':
            self._auto_range_axes()
        elif event.key == 'escape':
            self._clear_selection()
        elif event.key == 'z':
            self.undo()
        elif event.key in ('delete'):
            if self._canvas_has_focus():
                self._delete_selected_points()
        elif event.key == 'a':
            self._tab_index = 0
            self.set_active_curve("All")
        elif event.key == 'tab':
            if not self.radio_labels or len(self.curves) == 0:
                return
            self._tab_index = (self._tab_index + 1) % (len(self.curves) + 1)
            if self._tab_index == 0:
                self.set_active_curve("All")
            else:
                self.set_active_curve(self.curves[self._tab_index - 1]['name'])

    def save_curves_npy(self, event=None, filename="adjusted_curves.npy"):
        if not self.curves:
            return
        npy_data = np.array([curve['y'] for curve in self.curves]).T
        save_dir = ""
        if self.current_file_path:
            save_dir = splitext(self.current_file_path)[0] + "_"
        np.save(save_dir + filename, npy_data)
        print(f"NPY 已保存到: {save_dir + filename}  ({npy_data.shape[0]} 点 x {npy_data.shape[1]} 曲线)")

    def _reset_slider_widgets(self):
        self._slider_base_state = [{'y': crv['y'].copy()} for crv in self.curves]
        if hasattr(self, 'slider_scale') and self.slider_scale:
            self.slider_scale.eventson = False
            self.slider_scale.set_val(1.0)  
            self.slider_scale.eventson = True
        if hasattr(self, 'slider_smooth') and self.slider_smooth:
            self.slider_smooth.eventson = False
            self.slider_smooth.set_val(1)
            self.slider_smooth.eventson = True
        if hasattr(self, 'slider_noise') and self.slider_noise:
            self.slider_noise.eventson = False
            self.slider_noise.set_val(0.0)
            self.slider_noise.eventson = True

    def _get_target_points_map(self):
        effective = self._get_effective_active()
        targets = {}
        if self._selected_points:
            for c_idx, p_idx in self._selected_points:
                if c_idx in self.pinned_curve_indices:
                    continue
                if len(effective) == 0 or c_idx in effective:
                    targets.setdefault(c_idx, []).append(p_idx)
        elif len(effective) > 0:
            for c_idx in sorted(effective):
                if c_idx in self.pinned_curve_indices:
                    continue
                targets[c_idx] = list(range(len(self.curves[c_idx]['y'])))
        else:
            for c_idx in range(len(self.curves)):
                if c_idx in self.pinned_curve_indices:
                    continue
                targets[c_idx] = list(range(len(self.curves[c_idx]['y'])))
        return targets

    def on_slider_changed(self, val=None):
        if self._slider_base_state is None:
            return

        targets = self._get_target_points_map()
        scale_val = self.slider_scale.val
        smooth_val = int(self.slider_smooth.val)
        noise_coeff = self.slider_noise.val

        for c_idx, p_idxs in targets.items():
            orig_y = self._slider_base_state[c_idx]['y'].copy()
            p_idxs = np.array(p_idxs)
            if len(p_idxs) == 0:
                continue
            
            if scale_val != 1.0:
                orig_y[p_idxs] = orig_y[p_idxs] * scale_val

            if smooth_val > 1:
                smoothed = uniform_filter1d(orig_y.astype(float), size=smooth_val)
                orig_y[p_idxs] = smoothed[p_idxs]

            if noise_coeff > 0:
                data_range = np.std(orig_y[p_idxs]) if len(p_idxs) > 1 else np.mean(np.abs(orig_y[p_idxs]))
                if data_range == 0: data_range = 1.0

                rng = np.random.RandomState(self._noise_seed or 0)
                noise = rng.normal(0, data_range * 0.012 * noise_coeff, size=len(p_idxs))
                orig_y[p_idxs] += noise

            self.curves[c_idx]['y'][:] = orig_y
            self._update_spline(c_idx)

        self._expand_axes_if_needed()
        self._refresh_visual_style()
        self._update_selection_info()


# ===================== 演示数据 =====================
def _make_demo_curves(n_pts=80):
    """3 条干净利落的正弦曲线：低/中/高频，适合截图做 logo"""
    x = np.linspace(0, 3 * np.pi, n_pts)
    curves = [
        1.0  * np.sin(x),               # 基准波
        0.65 * np.sin(2.0 * x - 0.8),   # 倍频 + 小振幅
        0.45 * np.sin(3.5 * x + 0.5),   # 高频 + 更小振幅
    ]
    return np.column_stack(curves)  # (n_pts, 3)


# ===================== UI 工厂 =====================
def _build_ui():
    """创建 figure 和所有 axes，返回 (fig, axes_dict)"""
    # ── 清晰高对比度配色 ──
    BG_FIGURE  = '#dfe3e8'   # 整体背景 — 略深，衬托白色绘图区
    BG_PLOT    = '#fdfdfd'   # 主绘图区 — 近白
    GRID_COLOR = '#bcc3cc'   # 网格线 — 加深，结构感更强
    BG_SLIDER  = '#d5dbe3'   # 滑杆轨道底色
    BG_RADIO   = '#f2f4f7'   # 单选面板
    BG_INPUT   = '#e8ecf1'   # 输入框底色

    fig = plt.figure(figsize=(16, 9))
    fig.patch.set_facecolor(BG_FIGURE)
    fig.canvas.manager.set_window_title("CurveMagician")

    ax = {
        'main':     plt.axes([0.05, 0.10, 0.70, 0.80], facecolor=BG_PLOT),
        # 左侧操作按钮组：Open | Save | Undo | Fit
        'open':     plt.axes([0.05, 0.94, 0.08, 0.035]),
        'save':     plt.axes([0.14, 0.94, 0.08, 0.035]),
        'undo':     plt.axes([0.23, 0.94, 0.08, 0.035]),
        'fit':      plt.axes([0.32, 0.94, 0.06, 0.035]),
        'pin':      plt.axes([0.39, 0.94, 0.06, 0.035]),
        # 右侧参数组：+/-N | Points | Resample
        'neighbor': plt.axes([0.52, 0.94, 0.08, 0.035], facecolor=BG_INPUT),
        'npts':     plt.axes([0.61, 0.94, 0.07, 0.035], facecolor=BG_INPUT),
        'apply':    plt.axes([0.69, 0.94, 0.08, 0.035]),
        'scale':    plt.axes([0.80, 0.89, 0.18, 0.040], facecolor=BG_SLIDER),
        'smooth':   plt.axes([0.80, 0.84, 0.18, 0.040], facecolor=BG_SLIDER),
        'noise':    plt.axes([0.80, 0.79, 0.18, 0.040], facecolor=BG_SLIDER),
        'radio':    plt.axes([0.78, 0.08, 0.20, 0.68], facecolor=BG_RADIO),
    }

    # ── 主绘图区美化 ──
    main_ax = ax['main']
    main_ax.grid(True, linestyle='-', alpha=0.40, color=GRID_COLOR, linewidth=0.8)
    main_ax.tick_params(labelsize=9, colors='#3b4252')
    main_ax.spines['top'].set_visible(False)
    main_ax.spines['right'].set_visible(False)
    main_ax.spines['left'].set_color('#b0b8c4')
    main_ax.spines['bottom'].set_color('#b0b8c4')

    # ── 其他小 axes 统一隐藏边框 ──
    for key in ['open', 'save', 'undo', 'fit', 'pin', 'apply', 'neighbor', 'npts']:
        _hide_axes_spines(ax[key])

    return fig, ax


def _hide_axes_spines(axes):
    """隐藏 axes 的四个边框和刻度"""
    for spine in axes.spines.values():
        spine.set_visible(False)
    axes.set_xticks([])
    axes.set_yticks([])


def _wire_widgets(editor, ax):
    """创建按钮/滑杆/输入框并连接到编辑器"""
    # ── 按钮配色（高辨识度） ──
    C_OPEN   = '#2E86DE'  # 亮蓝
    C_SAVE   = '#10AC84'  # 翠绿
    C_UNDO   = '#EE5A24'  # 橘红
    C_FIT    = '#8E44AD'  # 明紫
    C_APPLY  = '#E44C3C'  # 朱红
    C_OPEN_H   = '#1e6fc0'
    C_SAVE_H   = '#0d8f6e'
    C_UNDO_H   = '#d04a1c'
    C_FIT_H    = '#753894'
    C_APPLY_H  = '#c83a2c'
    C_PIN      = '#00B894'  # 青绿
    C_PIN_H    = '#009976'

    # 左侧操作按钮组
    Button(ax['open'],  'Open',  color=C_OPEN,  hovercolor=C_OPEN_H).on_clicked(editor.on_open_clicked)
    Button(ax['save'],  'Save',  color=C_SAVE,  hovercolor=C_SAVE_H).on_clicked(editor.save_file)
    Button(ax['undo'],  'Undo',  color=C_UNDO,  hovercolor=C_UNDO_H).on_clicked(editor.undo)
    Button(ax['fit'],   'Fit',   color=C_FIT,   hovercolor=C_FIT_H).on_clicked(editor._auto_range_axes)
    Button(ax['pin'],   'Pin',   color=C_PIN,   hovercolor=C_PIN_H).on_clicked(editor.toggle_pin)

    # 右侧参数组
    TextBox(ax['neighbor'], "+/-N", initial="0").on_submit(editor.set_neighbor_num)

    resample_box = TextBox(ax['npts'], "", initial="")
    editor._resample_box = resample_box
    editor._typing_axes.add(ax['npts'])
    Button(ax['apply'], 'Resample', color=C_APPLY, hovercolor=C_APPLY_H).on_clicked(editor._do_resample)

    # ── 滑杆 ──
    SLIDER_COLOR = '#2E86DE'
    sliders = {
        'scale':  Slider(ax['scale'],  'Scale ',  0.0, 2.0, valinit=1.0, valfmt='%.2f', color=SLIDER_COLOR),
        'smooth': Slider(ax['smooth'], 'Smooth',  1,   15,  valinit=1,   valfmt='%d',   color=SLIDER_COLOR),
        'noise':  Slider(ax['noise'],  'Noise ',  0.0, 5.0, valinit=0.0, valfmt='%.1f', color=SLIDER_COLOR),
    }
    for s in sliders.values():
        s.label.set_size(10)
        s.label.set_color('#2d3436')
        s.label.set_fontweight('bold')
        s.valtext.set_fontsize(9)
        s.valtext.set_color('#2d3436')
        # 滑杆手柄（poly）美化
        if hasattr(s, 'poly'):
            s.poly.set_facecolor('#54a0ff')
            s.poly.set_edgecolor('#2E86DE')
            s.poly.set_linewidth(1.5)

    editor.slider_scale  = sliders['scale']
    editor.slider_smooth = sliders['smooth']
    editor.slider_noise  = sliders['noise']
    for s in sliders.values():
        s.on_changed(editor.on_slider_changed)


def _load_demo(editor, colors):
    """加载默认演示曲线"""
    data = _make_demo_curves()
    for idx in range(data.shape[1]):
        editor.add_curve(np.arange(data.shape[0]), data[:, idx],
                         color=colors[idx % len(colors)], name=f"Curve {idx}")
    editor._build_radio_panel(
        [colors[i % len(colors)] for i in range(data.shape[1])])
    editor._auto_range_axes()
    editor._save_current_state()
    editor._reset_slider_widgets()
    editor._update_selection_info()


# ===================== 主程序 =====================
if __name__ == "__main__":
    # 高饱和高区分度 15 色板 — 相邻曲线颜色差异大，看过去一目了然
    COLORS = ['#E44C3C', '#2E86DE', '#10AC84', '#EE5A24', '#8E44AD',
              '#F1C40F', '#1ABC9C', '#E84393', '#0984E3', '#00B894',
              '#6C5CE7', '#FD79A8', '#00CEC9', '#FF7675', '#A29BFE']

    fig, ax = _build_ui()
    editor = HarmoniousCurvesEditor(
        ax['main'], fig, ax['neighbor'], ax_radio=ax['radio'], color_list=COLORS)

    _wire_widgets(editor, ax)
    _load_demo(editor, COLORS)
    editor.setup_drag_and_drop()
    plt.show()