import matplotlib
matplotlib.use('TkAgg')

import numpy as np
import matplotlib.pyplot as plt
from os.path import splitext, basename
from matplotlib.patches import Rectangle, Circle
from matplotlib.widgets import AxesWidget, Button, TextBox, Slider
from scipy.interpolate import splprep, splev
from scipy.ndimage import uniform_filter1d
from tkinter import filedialog, messagebox

# 可选依赖：openpyxl 用于 Excel 读写（CSV 走 numpy）
try:
    from openpyxl import load_workbook, Workbook
    HAS_EXCEL = True
except ImportError:
    HAS_EXCEL = False
    print("提示: 安装 openpyxl 可支持 Excel 格式 (pip install openpyxl)")

# 可选依赖：tkinterdnd2 用于文件拖拽
try:
    from tkinterdnd2 import TkinterDnD, DND_FILES
    HAS_DND = True
except ImportError:
    HAS_DND = False
    print("提示: 安装 tkinterdnd2 可启用文件拖拽功能 (pip install tkinterdnd2)")

# ===================== 核心修复：精准、安全的 Matplotlib 补丁 =====================
def patch_matplotlib_widgets_safely():
    """
    精准修复 Matplotlib 组件基类的全局事件分发 Bug，增加防重复修改锁，彻底避免无限递归
    """
    base_class = AxesWidget
    
    # 检查是否已经修补过，防止重复执行
    if not getattr(base_class, '_is_patched_for_resize', False):
        orig_connect = base_class.connect_event
        
        def safe_connect(self, event, callback):
            def safe_callback(evt):
                if evt is None or not hasattr(evt, 'inaxes'):
                    return 
                return callback(evt)
            return orig_connect(self, event, safe_callback)
        
        base_class.connect_event = safe_connect
        base_class._is_patched_for_resize = True

# 启动补丁
patch_matplotlib_widgets_safely()


class HarmoniousCurvesEditor:
    def __init__(self, ax, fig, ax_n_input, ax_radio=None, color_list=None):
        self.fig = fig
        self.ax = ax
        self.canvas = ax.figure.canvas
        self.curves = []

        # 曲线激活锁定
        self.active_curve_idx = None
        self.current_label = "All"

        # 单点拖拽
        self._active_curve_idx = None
        self._active_point_idx = None

        # 增大容差像素，让鼠标更容易选中
        self._epsilon = 30

        self.neighbor_num = 0
        self.ax_n_input = ax_n_input
        self._typing_axes = {ax_n_input}  # 在这些区域内不响应 delete/backspace

        # 框选 & 批量移动
        self._selecting = False
        self._select_rect = None
        self._select_start = None
        self._selected_points = []
        self._batch_moving = False
        self._batch_origin = None
        self._selection_x_bounds = None

        # Undo 栈 & 临时操作状态备份
        self.undo_stack = []
        self._in_continuous_drag = False
        self._slider_base_state = None
        self._is_sliding = False  # 标记当前是否正在拖动滑杆
        self._noise_seed = None   # 单次拖拽缓存噪点种子，松手后重置
        self._tab_index = 0       # Tab 切换曲线索引，默认 All
        self._last_npts_val = None  # 上次同步到输入框的点数，避免覆盖用户输入

        # 组件引用初始化
        self.radio_menu = None
        self.radio_labels = []
        self.slider_scale = None
        self.slider_smooth = None
        self.slider_noise = None
        self._resample_box = None  # 外部注入的重采样输入框引用

        # 文件 I/O 状态
        self.current_file_path = None      # 当前文件路径（用于保存时推断格式）
        self.color_list = color_list or [] # 15 色调色板
        self.ax_radio = ax_radio           # 右侧单选面板 axes

        # 拖拽提示 & 单选面板顶层元素引用
        self.drop_hint = None              # 空数据时的拖拽提示文字
        self._top_dots = None              # 面板顶层圆点 scatter
        self._radio_circles = []           # 面板 Circle patch 列表
        self._color_patches = []           # 面板色块 Rectangle 列表
  
        # 底部状态信息 — 大号数字加粗，描述浅色跟随
        self.info_num = self.fig.text(
            0.05, 0.02, "", fontsize=15, fontweight='bold',
            color='#2e3440', va='bottom', ha='left',
        )
        self.info_tag = self.fig.text(
            0.15, 0.02, "", fontsize=10, color='#4c566a',
            fontweight='normal', va='bottom', ha='left',
        )

        # 事件绑定
        self.canvas.mpl_connect('button_press_event', self.on_press)
        self.canvas.mpl_connect('motion_notify_event', self.on_motion)
        self.canvas.mpl_connect('button_release_event', self.on_release)
        self.canvas.mpl_connect('key_press_event', self.on_key)
        self.canvas.mpl_connect('button_press_event', self.on_double_click)

        # Tk 级别绑定 Tab（matplotlib key_press_event 收不到 Tab）
        self._setup_tk_tab()
        # 点击绘图区时确保键盘焦点回到 canvas
        self.canvas.mpl_connect('button_press_event', self._ensure_focus)

        self._update_selection_info()

    def set_radio_ref(self, radio_menu, labels):
        self.radio_menu = radio_menu
        self.radio_labels = labels

    def _is_toolbar_active(self):
        toolbar = self.fig.canvas.manager.toolbar
        if toolbar is not None:
            mode = toolbar.mode.lower()
            if 'zoom' in mode or 'pan' in mode:
                return True
        return False

    def on_double_click(self, event):
        if not hasattr(event, 'inaxes') or event.inaxes is None:
            return
        if event.button != 1 or not event.dblclick:
            return
        if event.inaxes != self.ax or self._is_toolbar_active():
            return
        c_idx, _ = self._get_closest_point(event)
        if c_idx is None:
            return
        target_label = f"Curve {c_idx}"
        if self.radio_menu and target_label in self.radio_labels:
            self.radio_menu.set_active(self.radio_labels.index(target_label))
            self.set_active_curve(target_label)

    def set_active_curve(self, label):
        self.current_label = label
        if label == "All":
            self.active_curve_idx = None
        else:
            for idx, curve in enumerate(self.curves):
                if curve['name'] == label:
                    self.active_curve_idx = idx
        if self._selected_points and label != "All":
            valid_points = [(c, p) for c, p in self._selected_points if c == self.active_curve_idx]
            self._selected_points = valid_points
        self._refresh_visual_style()
        self._reset_slider_widgets()
        self._update_selection_info()

    def _refresh_visual_style(self):
        for idx, curve in enumerate(self.curves):
            is_curve_active = (self.active_curve_idx is None or idx == self.active_curve_idx)
            if self._selection_x_bounds is not None:
                curve['ctrl_points'].set_alpha(0.02)
                curve['spline_line'].set_alpha(0.1)
                curve['spline_line'].set_linewidth(0.8)
                has_points_selected = any(c == idx for c, _ in self._selected_points)
                if is_curve_active and has_points_selected:
                    sel_p_idxs = [p for c, p in self._selected_points if c == idx]
                    xs_p = curve['x'][sel_p_idxs]
                    ys_p = curve['y'][sel_p_idxs]
                    curve['hl_ctrl_points'].set_data(xs_p, ys_p)
                    curve['hl_ctrl_points'].set_alpha(1.0)
                    curve['hl_ctrl_points'].set_markersize(7)
                    x_fine, y_fine = curve['fine_x_full'], curve['fine_y_full']
                    xmin, xmax = self._selection_x_bounds
                    mask = (x_fine >= xmin) & (x_fine <= xmax)
                    if np.any(mask):
                        curve['hl_spline_line'].set_data(x_fine[mask], y_fine[mask])
                        curve['hl_spline_line'].set_alpha(1.0)
                        curve['hl_spline_line'].set_linewidth(4.0)
                    else:
                        curve['hl_spline_line'].set_data([], [])
                else:
                    curve['hl_spline_line'].set_data([], [])
                    curve['hl_ctrl_points'].set_data([], [])
            else:
                curve['hl_spline_line'].set_data([], [])
                curve['hl_ctrl_points'].set_data([], [])
                if is_curve_active:
                    alpha_ctrl = 0.3 if self.active_curve_idx is None else 0.8
                    lw = 2.0 if self.active_curve_idx is None else 3.5
                    ms = 4 if self.active_curve_idx is None else 6
                    curve['ctrl_points'].set_alpha(alpha_ctrl)
                    curve['ctrl_points'].set_markersize(ms)
                    curve['spline_line'].set_alpha(1.0)
                    curve['spline_line'].set_linewidth(lw)
                else:
                    curve['ctrl_points'].set_alpha(0.05)
                    curve['ctrl_points'].set_markersize(3)
                    curve['spline_line'].set_alpha(0.15)
                    curve['spline_line'].set_linewidth(1.0)
        self.canvas.draw_idle()

    def _update_selection_info(self):
        """显示当前选中的控制点信息，并实时初始化 N-pts 右侧重采样输入框"""
        if not self.curves:
            self.info_num.set_text("")
            self.info_tag.set_text("")
            return

        current_resample_val = ""

        if self._selected_points:
            # 1. 处于框选状态：统计每条曲线被选中的点数
            by_curve = {}
            for c, p in self._selected_points:
                by_curve.setdefault(c, []).append(p)
            
            n = len(self._selected_points)
            detail = ", ".join(f"C{c}: {len(p_)}p" for c, p_ in sorted(by_curve.items()))
            self.info_num.set_text(f"{n} pts")
            self.info_tag.set_text(f"Box-sel  ({detail})")

            # 提取所有被选中曲线片段的点数集合
            selected_counts = {len(p_idxs) for p_idxs in by_curve.values()}
            if len(selected_counts) == 1:
                # 只有当框选的每条曲线片段内的点数完全一致时，才在输入框初始化该数字
                current_resample_val = str(selected_counts.pop())
            else:
                current_resample_val = "---"  # 不一致则显示占位符，提示输入

        elif self.active_curve_idx is not None:
            # 2. 未框选，但单选了某一条曲线
            n = len(self.curves[self.active_curve_idx]['y'])
            self.info_num.set_text(f"{n} pts")
            self.info_tag.set_text(f"{self.current_label}")
            current_resample_val = str(n)
        else:
            # 3. 默认 All 状态，未进行框选
            total = sum(len(c['y']) for c in self.curves)
            self.info_num.set_text(f"{total} pts")
            nc, ppc = len(self.curves), len(self.curves[0]['y'])
            self.info_tag.set_text(f"All ({nc} curves x {ppc})")
            
            # 检查是否所有曲线的总长度都相同
            all_counts = {len(c['y']) for c in self.curves}
            current_resample_val = str(all_counts.pop()) if len(all_counts) == 1 else "---"

        # 同步输入框：点数一致填数字，不一致填 "---"
        if hasattr(self, '_resample_box') and self._resample_box:
            if current_resample_val != self._last_npts_val:
                self._resample_box.set_val(current_resample_val)
                self._last_npts_val = current_resample_val

        self.canvas.draw_idle()

    def _canvas_has_focus(self):
        try:
            canvas = self.fig.canvas.get_tk_widget()
            focused = canvas.focus_get()
            return focused is None or focused is canvas
        except Exception:
            return True

    def _setup_tk_tab(self):
        try:
            canvas_widget = self.fig.canvas.get_tk_widget()
            def _on_tk_tab(event):
                if not self.radio_labels:
                    return
                self._tab_index = (self._tab_index + 1) % len(self.radio_labels)
                if self.radio_menu:
                    self.radio_menu.set_active(self._tab_index)
                self.set_active_curve(self.radio_labels[self._tab_index])
                return 'break'
            canvas_widget.bind('<Tab>', _on_tk_tab)
        except Exception:
            pass

    def _ensure_focus(self, event=None):
        try:
            self.fig.canvas.get_tk_widget().focus_set()
        except Exception:
            pass

    def _clear_selection(self):
        self._selected_points.clear()
        self._batch_moving = False
        self._selection_x_bounds = None
        if self._select_rect is not None:
            self._select_rect.remove()
            self._select_rect = None
        self._refresh_visual_style()
        self._reset_slider_widgets()
        self._update_selection_info()

    def _delete_selected_points(self):
        if not self._selected_points:
            return
        by_curve = {}
        for c, p in self._selected_points:
            by_curve.setdefault(c, []).append(p)

        deleted_any = False
        for c_idx, p_idxs in by_curve.items():
            curve = self.curves[c_idx]
            keep = [i for i in range(len(curve['x'])) if i not in p_idxs]
            if len(keep) < 4:
                print(f"  {curve['name']} 删除后不足 4 点，跳过")
                continue
            curve['x'] = np.arange(len(keep), dtype=float)
            curve['y'] = curve['y'][keep]
            self._update_spline(c_idx)
            deleted_any = True

        if deleted_any:
            self._save_current_state()
        self._clear_selection()

    def _do_resample(self, event=None):
        """核心业务功能：将选中曲线的局部控制点，平滑增点/降点重采样为 Y 个点"""
        if not hasattr(self, '_resample_box') or not self._resample_box:
            return
        
        try:
            n_new = int(self._resample_box.text)
        except ValueError:
            messagebox.showwarning("输入错误", "请输入有效的段内目标点数整数（如 10）")
            return
        if n_new < 4:
            messagebox.showwarning("约束错误", "局部段重采样后至少需要保留 4 个控制点以维持三次样条拟合")
            return

        # 1. 确定当前哪些曲线和哪些点将被作为目标处理
        by_curve = {}
        if self._selected_points:
            for c, p in self._selected_points:
                if self.active_curve_idx is None or c == self.active_curve_idx:
                    by_curve.setdefault(c, []).append(p)
        else:
            targets = [self.active_curve_idx] if self.active_curve_idx is not None else list(range(len(self.curves)))
            for c in targets:
                by_curve[c] = list(range(len(self.curves[c]['y'])))

        if not by_curve:
            print("当前未选中任何有效曲线或控制点")
            return

        # 2. 【强约束判定】确保当前所有选中的曲线，它们被选中的点数（X）必须完全一致
        counts = {len(p_idxs) for p_idxs in by_curve.values()}
        if len(counts) > 1:
            messagebox.showerror("重采样失败", f"强约束未满足！当前框选的各曲线内部点数不一致: {counts}，无法批量重采样。")
            return

        modified_any = False

        # 3. 开始执行带边界保护的三次样条局部重采样
        for c_idx, p_idxs in by_curve.items():
            curve = self.curves[c_idx]
            orig_y = curve['y']
            total_pts = len(orig_y)

            i_min, i_max = min(p_idxs), max(p_idxs)
            n_old = i_max - i_min + 1

            if n_old == n_new and self._selected_points:
                continue  # 数量一样且是局部框选，不作变动

            # 为保证交接处极度平滑，向左右未选中区外扩 2 个控制点作为“缓冲约束区”
            pad = 2
            ext_min = max(0, i_min - pad)
            ext_max = min(total_pts - 1, i_max + pad)

            seg_y = orig_y[ext_min:ext_max + 1]
            seg_x = np.arange(len(seg_y), dtype=float)

            k_order = 3 if len(seg_y) > 3 else (len(seg_y) - 1)
            if k_order < 1:
                continue

            # 构建样条核心参数
            tck, u = splprep([seg_x, seg_y], s=0, k=k_order)

            # 计算生成的新段里包含的缓冲区节点数
            left_pad_count = i_min - ext_min
            right_pad_count = ext_max - i_max
            total_seg_new_count = left_pad_count + n_new + right_pad_count

            u_new = np.linspace(0, 1, total_seg_new_count)
            _, y_resampled_ext = splev(u_new, tck)

            # 剔除缓冲区，完美截取并恢复我们需要的 Y 个新点
            start_idx = left_pad_count
            end_idx = total_seg_new_count - right_pad_count
            y_resampled_core = y_resampled_ext[start_idx:end_idx]

            # 完美拼回大数组
            y_new = np.concatenate([
                orig_y[:i_min],
                y_resampled_core,
                orig_y[i_max + 1:]
            ])

            # 重新构建自增一维 X 轴索引并更新艺术家图元
            curve['x'] = np.arange(len(y_new), dtype=float)
            curve['y'] = y_new
            self._update_spline(c_idx)
            modified_any = True

        if modified_any:
            self._save_current_state()
            print(f"成功！已将各曲线选中段的控制点平滑转换为了 {n_new} 个点。")
        
        # 联动重置清除框选状态
        self._clear_selection()

    def add_curve(self, x_ctrl, y_ctrl, color='blue', name="Curve"):
        x = np.array(x_ctrl, dtype=float)
        y = np.array(y_ctrl, dtype=float)
        if len(x) < 4:
            raise ValueError("At least 4 control points required.")
        ctrl_points, = self.ax.plot(x, y, 'o', color=color, markersize=4, alpha=0.4, zorder=2)
        spline_line, = self.ax.plot([], [], '-', color=color, linewidth=2, label=name, zorder=1)
        hl_ctrl_points, = self.ax.plot([], [], 'o', color=color, markersize=7, alpha=0.0, zorder=4)
        hl_spline_line, = self.ax.plot([], [], '-', color=color, linewidth=4, alpha=0.0, zorder=3)
        curve_dict = {
            'x': x, 'y': y,
            'color': color,
            'ctrl_points': ctrl_points,
            'spline_line': spline_line,
            'hl_ctrl_points': hl_ctrl_points,
            'hl_spline_line': hl_spline_line,
            'fine_x_full': None, 'fine_y_full': None,
            'name': name
        }
        self.curves.append(curve_dict)
        self._update_spline(len(self.curves) - 1)

    # ===================== 文件 I/O =====================

    def _clear_all_curves(self):
        for curve in self.curves:
            for key in ['ctrl_points', 'spline_line', 'hl_ctrl_points', 'hl_spline_line']:
                if curve.get(key) is not None:
                    curve[key].remove()
        self.curves.clear()
        self.undo_stack.clear()
        self._selected_points.clear()
        self._batch_moving = False
        self._selection_x_bounds = None
        self._slider_base_state = None
        self.active_curve_idx = None
        self.current_label = "All"
        self._active_curve_idx = None
        self._active_point_idx = None
        self._tab_index = 0
        self._last_npts_val = None
        self._hide_drop_hint()
        self.ax.relim()
        self.ax.autoscale_view()

    def _auto_range_axes(self, event=None):
        if not self.curves:
            return
        all_x = np.concatenate([c['x'] for c in self.curves])
        all_y = np.concatenate([c['y'] for c in self.curves])
        if len(all_x) == 0 or len(all_y) == 0:
            return

        x_min, x_max = np.min(all_x), np.max(all_x)
        y_min, y_max = np.min(all_y), np.max(all_y)

        x_range = x_max - x_min or 1.0
        y_range = y_max - y_min or 1.0
        margin = 0.08

        self.ax.set_xlim(x_min - x_range * margin, x_max + x_range * margin)
        self.ax.set_ylim(y_min - y_range * margin, y_max + y_range * margin)
        self.canvas.draw_idle()

    def _expand_axes_if_needed(self):
        if not self.curves:
            return
        all_y = np.concatenate([c['y'] for c in self.curves])
        if len(all_y) == 0:
            return

        y_min, y_max = np.min(all_y), np.max(all_y)
        cur_ylo, cur_yhi = self.ax.get_ylim()

        y_range = y_max - y_min or 1.0
        margin = 0.10

        new_lo, new_hi = cur_ylo, cur_yhi
        changed = False

        if y_min < cur_ylo:
            new_lo = y_min - y_range * margin
            changed = True
        if y_max > cur_yhi:
            new_hi = y_max + y_range * margin
            changed = True

        if changed:
            self.ax.set_ylim(new_lo, new_hi)
            self.canvas.draw_idle()

    def _detect_format(self, file_path):
        ext = splitext(file_path)[1].lower()
        if ext == '.csv':
            return 'csv'
        elif ext in ('.xlsx', '.xls'):
            return 'excel'
        elif ext == '.npy':
            return 'npy'
        else:
            raise ValueError(f"不支持的文件格式: {ext}（支持 .csv .xlsx .npy）")

    def _load_data_from_file(self, file_path):
        fmt = self._detect_format(file_path)

        if fmt == 'csv':
            data = np.loadtxt(file_path, delimiter=',', dtype=float, ndmin=2)
        elif fmt == 'excel':
            if not HAS_EXCEL:
                raise ImportError("读取 Excel 需要 openpyxl: pip install openpyxl")
            wb = load_workbook(file_path, data_only=True)
            ws = wb.active
            rows = [[cell.value or 0 for cell in row] for row in ws.iter_rows()]
            if not rows:
                raise ValueError("Excel 文件为空")
            data = np.array(rows, dtype=float)
            wb.close()
        elif fmt == 'npy':
            data = np.load(file_path, allow_pickle=True)

        if data.ndim == 1:
            data = data.reshape(-1, 1)
        if data.ndim != 2:
            raise ValueError(f"数据必须是 1D 或 2D，当前 shape: {data.shape}")

        n_rows, n_cols = data.shape
        if n_rows < n_cols:
            data = data.T
            print(f"  方向检测: {n_rows}行×{n_cols}列 → 自动转置为 {n_cols}点×{n_rows}曲线")
        elif n_cols < n_rows:
            print(f"  方向检测: {n_rows}行×{n_cols}列 → 已是列式 ({n_rows}点×{n_cols}曲线)")

        if data.shape[0] < 4:
            raise ValueError(f"每条曲线至少需要 4 个采样点，当前仅 {data.shape[0]} 行")
        return data

    def load_curves_from_file(self, file_path=None):
        if file_path is None:
            file_path = filedialog.askopenfilename(
                title="打开曲线文件",
                filetypes=[
                    ("所有支持格式", "*.csv;*.xlsx;*.xls;*.npy"),
                    ("CSV 文件", "*.csv"),
                    ("Excel 文件", "*.xlsx;*.xls"),
                    ("NumPy 文件", "*.npy"),
                ]
            )
            if not file_path:
                return

        try:
            data = self._load_data_from_file(file_path)
        except Exception as e:
            messagebox.showerror("加载错误", f"无法加载文件:\n{e}")
            return

        self.current_file_path = file_path
        self._clear_all_curves()
        self._clear_radio_panel()

        n_curves = data.shape[1]
        n_points = data.shape[0]
        curve_colors = []
        for idx in range(n_curves):
            color = self.color_list[idx % len(self.color_list)] if self.color_list else f"C{idx}"
            name = f"Curve {idx}"
            curve_colors.append(color)
            self.add_curve(np.arange(n_points), data[:, idx], color=color, name=name)

        self._build_radio_panel(curve_colors)
        self._auto_range_axes()
        self._save_current_state()
        self._reset_slider_widgets()
        self._hide_drop_hint()
        self._update_selection_info()

        fname = basename(file_path)
        self.fig.canvas.manager.set_window_title(f"CurveMagician - {fname}")
        print(f"已加载 {n_curves} 条曲线 ({n_points} 点/条)，来自 {file_path}")

    def save_file(self, event=None):
        if not self.curves:
            messagebox.showwarning("保存", "没有曲线数据可保存。")
            return

        default_ext = ".npy"
        initial_file = "curves.npy"
        if self.current_file_path:
            default_ext = splitext(self.current_file_path)[1]
            initial_file = basename(self.current_file_path)

        file_path = filedialog.asksaveasfilename(
            title="保存曲线为",
            initialfile=initial_file,
            defaultextension=default_ext,
            filetypes=[
                ("NumPy 文件", "*.npy"),
                ("CSV 文件", "*.csv"),
                ("Excel 文件", "*.xlsx"),
            ]
        )
        if not file_path:
            return

        try:
            fmt = self._detect_format(file_path)
            data = np.array([curve['y'] for curve in self.curves]).T

            if fmt == 'csv':
                np.savetxt(file_path, data, delimiter=',', fmt='%.8g')
            elif fmt == 'excel':
                if not HAS_EXCEL:
                    raise ImportError("保存 Excel 需要 openpyxl: pip install openpyxl")
                wb = Workbook()
                ws = wb.active
                for row in data:
                    ws.append(row.tolist())
                wb.save(file_path)
            elif fmt == 'npy':
                np.save(file_path, data)

            self.current_file_path = file_path
            fname = basename(file_path)
            self.fig.canvas.manager.set_window_title(f"CurveMagician - {fname}")
            print(f"已保存 {data.shape[1]} 条曲线 ({data.shape[0]} 点/条) 到 {file_path}")
        except Exception as e:
            messagebox.showerror("保存错误", f"保存失败:\n{e}")

    def on_open_clicked(self, event=None):
        self.load_curves_from_file()

    # ===================== 单选面板构建 =====================

    def _clear_radio_panel(self):
        if self.ax_radio is None:
            return
        self.ax_radio.cla()
        self.ax_radio.set_facecolor('#fafbfc')
        self.ax_radio.set_xticks([])
        self.ax_radio.set_yticks([])
        self.ax_radio.set_navigate(False)
        self.ax_radio.set_xlim(0, 1)
        self.ax_radio.set_ylim(0, 1)
        self._top_dots = None
        self._radio_circles.clear()
        self._color_patches.clear()
        self.radio_menu = None
        self.radio_labels = []

    def _build_radio_panel(self, curve_colors):
        if self.ax_radio is None:
            return

        ax = self.ax_radio
        ax.set_facecolor('#fafbfc')
        base_labels = ["All"] + [self.curves[i]['name'] for i in range(len(curve_colors))]
        num_labels = len(base_labels)
        font_size = 9 if num_labels > 8 else 10
        activecolor = '#5e81ac'   # 选中态钢蓝
        active_idx = [0]

        ys = np.linspace(1, 0, num_labels + 2)[1:-1]
        dot_radius = 0.022 if num_labels > 8 else 0.030
        patch_h = 0.55 / num_labels if num_labels > 8 else 0.038

        # ── 面板标题 ──
        ax.text(0.50, 0.97, 'Curves', transform=ax.transAxes,
                fontsize=11, fontweight='bold', color='#4c566a',
                ha='center', va='top')

        self._radio_circles = []

        for i in range(num_labels):
            y = ys[i]

            # 圆点 — 空心 + 选中态填充
            c = Circle(
                (0.09, y), dot_radius,
                transform=ax.transAxes,
                facecolor='white',
                edgecolor='#a0a8b8',
                linewidth=1.8,
                zorder=10,
            )
            ax.add_patch(c)
            self._radio_circles.append(c)

            # 曲线名称
            ax.text(
                0.42, y, base_labels[i],
                transform=ax.transAxes,
                fontsize=font_size,
                fontweight='bold' if i == 0 else 'normal',
                color='#4c566a',
                va='center',
            )

            if i == 0:
                continue

            # 色块 — 圆角视觉 + 柔和边框
            rect = Rectangle(
                (0.18, y - patch_h / 2), 0.20, patch_h,
                facecolor=curve_colors[i - 1],
                transform=ax.transAxes,
                zorder=0,
                edgecolor='#d8dee9',
                linewidth=0.8,
                alpha=0.85,
            )
            ax.add_patch(rect)
            self._color_patches.append(rect)

        def _refresh_dots():
            for i, c in enumerate(self._radio_circles):
                if i == active_idx[0]:
                    c.set_facecolor(activecolor)
                    c.set_edgecolor('#4c6f99')
                    c.set_linewidth(2.2)
                else:
                    c.set_facecolor('white')
                    c.set_edgecolor('#a0a8b8')
                    c.set_linewidth(1.8)

        _refresh_dots()

        def _on_radio_click(event):
            if event.inaxes != ax or event.button != 1:
                return
            if event.ydata is None:
                return
            dists = np.abs(ys - event.ydata)
            closest = int(np.argmin(dists))
            if dists[closest] < 0.8 / num_labels:
                active_idx[0] = closest
                _refresh_dots()
                self.fig.canvas.draw_idle()
                self.set_active_curve(base_labels[closest])

        self.fig.canvas.mpl_connect('button_press_event', _on_radio_click)

        class _RadioCompat:
            labels = base_labels

            @staticmethod
            def set_active(idx):
                active_idx[0] = idx
                _refresh_dots()
                self.fig.canvas.draw_idle()

            @staticmethod
            def on_clicked(_cb):
                pass

        self.radio_menu = _RadioCompat()
        self.radio_labels = base_labels

    # ===================== 拖拽支持 =====================

    def _show_drop_hint(self):
        if self.drop_hint is None:
            self.drop_hint = self.ax.text(
                0.5, 0.5,
                '拖放 CSV / Excel / NPY 文件到此处\n或点击 Open 按钮选择文件',
                transform=self.ax.transAxes,
                ha='center', va='center',
                fontsize=15, color='#4c566a', alpha=0.30,
                fontweight='bold',
                zorder=100,
            )
        else:
            self.drop_hint.set_visible(True)
        self.canvas.draw_idle()

    def _hide_drop_hint(self):
        if self.drop_hint is not None:
            self.drop_hint.set_visible(False)
            self.canvas.draw_idle()

    def _on_drop(self, event):
        files = self.fig.canvas.manager.window.tk.splitlist(event.data)
        if not files:
            return

        cleaned = []
        for f in files:
            f = f.strip()
            if f.startswith('{') and f.endswith('}'):
                f = f[1:-1]
            cleaned.append(f)

        for fpath in cleaned:
            try:
                self._detect_format(fpath)
                self.load_curves_from_file(fpath)
                return
            except ValueError:
                continue

        messagebox.showwarning(
            "不支持的文件",
            "未找到支持的文件格式。\n支持: .csv, .xlsx, .npy"
        )

    def setup_drag_and_drop(self):
        if not HAS_DND:
            return
        try:
            import tkinter as tk
            root = tk._default_root or self.fig.canvas.manager.window.winfo_toplevel()
            TkinterDnD.require(root)
            canvas_widget = self.fig.canvas.get_tk_widget()
            canvas_widget.drop_target_register(DND_FILES)
            canvas_widget.dnd_bind('<<Drop>>', self._on_drop)
            print("文件拖拽功能已启用")
        except Exception as e:
            print(f"拖拽注册失败: {e}")

    def _update_spline(self, curve_idx):
        curve = self.curves[curve_idx]
        x, y = curve['x'], curve['y']
        tck, u = splprep([x, y], s=0, k=3)
        u_fine = np.linspace(0, 1, 300)
        x_fine, y_fine = splev(u_fine, tck)
        curve['ctrl_points'].set_data(x, y)
        curve['spline_line'].set_data(x_fine, y_fine)
        curve['fine_x_full'] = x_fine
        curve['fine_y_full'] = y_fine

    def on_press(self, event):
        if self._is_toolbar_active():
            return
        
        if not hasattr(event, 'inaxes') or event.inaxes is None:
            return

        slider_axes = []
        if self.slider_scale: slider_axes.append(self.slider_scale.ax)
        if self.slider_smooth: slider_axes.append(self.slider_smooth.ax)
        if self.slider_noise: slider_axes.append(self.slider_noise.ax)

        if event.inaxes in slider_axes:
            self._is_sliding = True
            self._noise_seed = np.random.randint(0, 2**31)
            return

        if event.inaxes != self.ax:
            return

        if event.button == 3:
            self._clear_selection()
            return
        if event.button == 1:
            if self._selected_points:
                self._batch_moving = True
                self._batch_origin = (event.xdata, event.ydata)
                self._in_continuous_drag = True
                return
            c_idx, p_idx = self._get_closest_point(event)
            if c_idx is not None and p_idx is not None:
                self._active_curve_idx = c_idx
                self._active_point_idx = p_idx
                self._in_continuous_drag = True
                self.curves[c_idx]['ctrl_points'].set_markersize(8)
                self.curves[c_idx]['ctrl_points'].set_alpha(1.0)
                self.canvas.draw_idle()
                return
            self._selecting = True
            self._select_start = (event.xdata, event.ydata)
            self._select_rect = Rectangle(self._select_start, 0, 0, fill=True, alpha=0.2, color='gray')
            self.ax.add_patch(self._select_rect)

    def on_motion(self, event):
        if self._is_toolbar_active():
            return
        if self._is_sliding:
            return  

        if not hasattr(event, 'inaxes') or event.inaxes != self.ax or event.ydata is None:
            return

        if self._active_curve_idx is not None:
            c_idx = self._active_curve_idx
            p_center = self._active_point_idx
            curve = self.curves[c_idx]
            pts_idx = self._get_linked_point_indices(len(curve['y']), p_center)
            dy = event.ydata - curve['y'][p_center]
            for pid in pts_idx:
                curve['y'][pid] += dy
            self._update_spline(c_idx)
            if self._selection_x_bounds is not None:
                self._refresh_visual_style()
            else:
                self.canvas.draw_idle()
            return
        if self._selecting and self._select_rect is not None:
            x0, y0 = self._select_start
            x1, y1 = event.xdata, event.ydata
            self._select_rect.set_width(x1 - x0)
            self._select_rect.set_height(y1 - y0)
            self.canvas.draw_idle()
            return
        if self._batch_moving and self._batch_origin is not None:
            dy = event.ydata - self._batch_origin[1]
            for c_idx, p_idx in self._selected_points:
                if self.active_curve_idx is None or c_idx == self.active_curve_idx:
                    self.curves[c_idx]['y'][p_idx] += dy
            affected = set(c for c, _ in self._selected_points)
            for c in affected:
                self._update_spline(c)
            self._batch_origin = (self._batch_origin[0], event.ydata)
            self._refresh_visual_style()

    def on_release(self, event):
        if self._selecting:
            self._selecting = False
            if (self._select_start and event.xdata is not None and event.ydata is not None):
                x0, y0 = self._select_start
                x1, y1 = event.xdata, event.ydata
                self._selection_x_bounds = (min(x0, x1), max(x0, x1))
                self._selected_points = self._select_points_in_rect(x0, y0, x1, y1)
                self._refresh_visual_style()
            if self._select_rect is not None:
                self._select_rect.remove()
                self._select_rect = None
            self._update_selection_info()

        if self._is_toolbar_active():
            return

        if self._is_sliding:
            self._is_sliding = False
            self._noise_seed = None
            self._save_current_state()
            self._slider_base_state = [{'y': crv['y'].copy()} for crv in self.curves]
            return

        if not hasattr(event, 'inaxes') or event.inaxes is None:
            return

        need_save = False
        if self._active_curve_idx is not None:
            need_save = True
            self._refresh_visual_style()

        if self._batch_moving:
            need_save = True

        self._active_curve_idx = None
        self._active_point_idx = None
        self._batch_moving = False
        self._batch_origin = None
        self._in_continuous_drag = False

        if need_save:
            self._save_current_state()
            self._reset_slider_widgets()
            self._expand_axes_if_needed()

        self._update_selection_info()

    def _get_closest_point(self, event):
        if not self.curves or event.xdata is None or event.ydata is None:
            return None, None
        mouse_xy = self.ax.transData.transform((event.xdata, event.ydata))
        if self.active_curve_idx is not None:
            c_idx = self.active_curve_idx
            curve = self.curves[c_idx]
            x, y = curve['x'], curve['y']
            pts_xy = self.ax.transData.transform(np.c_[x, y])
            distances = np.linalg.norm(pts_xy - mouse_xy, axis=1)
            p_idx = np.argmin(distances)
            if distances[p_idx] < self._epsilon:
                return c_idx, p_idx
        else:
            min_dist = float('inf')
            closest_curve, closest_point = None, None
            for c_idx, curve in enumerate(self.curves):
                x, y = curve['x'], curve['y']
                pts_xy = self.ax.transData.transform(np.c_[x, y])
                distances = np.linalg.norm(pts_xy - mouse_xy, axis=1)
                p_idx = np.argmin(distances)
                if distances[p_idx] < min_dist:
                    min_dist = distances[p_idx]
                    closest_curve, closest_point = c_idx, p_idx
            if min_dist < self._epsilon:
                return closest_curve, closest_point
        return None, None

    def _select_points_in_rect(self, x0, y0, x1, y1):
        selected = []
        xmin, xmax = min(x0, x1), max(x0, x1)
        ymin, ymax = min(y0, y1), max(y0, y1)
        curves_to_check = [(self.active_curve_idx, self.curves[self.active_curve_idx])] if self.active_curve_idx is not None else enumerate(self.curves)
        for c_idx, curve in curves_to_check:
            xs, ys = curve['x'], curve['y']
            for p_idx, (px, py) in enumerate(zip(xs, ys)):
                if xmin <= px <= xmax and ymin <= py <= ymax:
                    selected.append((c_idx, p_idx))
        return selected

    def _get_linked_point_indices(self, total_len, center_idx):
        n = self.neighbor_num
        return list(range(max(0, center_idx - n), min(total_len - 1, center_idx + n) + 1))

    def set_neighbor_num(self, text):
        try:
            val = int(text)
            self.neighbor_num = max(0, val)
            self._update_selection_info()
        except ValueError:
            print("Please enter an integer!")

    def _save_current_state(self):
        state = [{'x': crv['x'].copy(), 'y': crv['y'].copy()} for crv in self.curves]
        self.undo_stack.append(state)

    def undo(self, event=None):
        if len(self.undo_stack) <= 1:
            return
        self.undo_stack.pop()
        prev_state = self.undo_stack[-1]
        for idx, crv in enumerate(self.curves):
            if idx < len(prev_state):
                crv['x'] = prev_state[idx]['x'].copy()
                crv['y'] = prev_state[idx]['y'].copy()
                self._update_spline(idx)
        self._refresh_visual_style()
        self._reset_slider_widgets()

    def on_key(self, event):
        if event.key == 's':
            self.save_file()
        elif event.key == 'o':
            self.on_open_clicked()
        elif event.key == 'r':
            self._auto_range_axes()
        elif event.key == 'escape':
            self._clear_selection()
        elif event.key == 'z':
            self.undo()
        elif event.key in ('delete'):
            if self._canvas_has_focus():
                self._delete_selected_points()
        elif event.key == 'a':
            self._tab_index = 0
            if self.radio_menu and self.radio_labels:
                self.radio_menu.set_active(0)
            self.set_active_curve("All")
        elif event.key == 'tab':
            if not self.radio_labels:
                return
            self._tab_index = (self._tab_index + 1) % len(self.radio_labels)
            if self.radio_menu:
                self.radio_menu.set_active(self._tab_index)
            self.set_active_curve(self.radio_labels[self._tab_index])

    def save_curves_npy(self, event=None, filename="adjusted_curves.npy"):
        if not self.curves:
            return
        npy_data = np.array([curve['y'] for curve in self.curves]).T
        save_dir = ""
        if self.current_file_path:
            save_dir = splitext(self.current_file_path)[0] + "_"
        np.save(save_dir + filename, npy_data)
        print(f"NPY 已保存到: {save_dir + filename}  ({npy_data.shape[0]} 点 x {npy_data.shape[1]} 曲线)")

    def _reset_slider_widgets(self):
        self._slider_base_state = [{'y': crv['y'].copy()} for crv in self.curves]
        if hasattr(self, 'slider_scale') and self.slider_scale:
            self.slider_scale.eventson = False
            self.slider_scale.set_val(1.0)  
            self.slider_scale.eventson = True
        if hasattr(self, 'slider_smooth') and self.slider_smooth:
            self.slider_smooth.eventson = False
            self.slider_smooth.set_val(1)
            self.slider_smooth.eventson = True
        if hasattr(self, 'slider_noise') and self.slider_noise:
            self.slider_noise.eventson = False
            self.slider_noise.set_val(0.0)
            self.slider_noise.eventson = True

    def _get_target_points_map(self):
        targets = {}
        if self._selected_points:
            for c_idx, p_idx in self._selected_points:
                if self.active_curve_idx is None or c_idx == self.active_curve_idx:
                    targets.setdefault(c_idx, []).append(p_idx)
        elif self.active_curve_idx is not None:
            targets[self.active_curve_idx] = list(range(len(self.curves[self.active_curve_idx]['y'])))
        else:
            for c_idx in range(len(self.curves)):
                targets[c_idx] = list(range(len(self.curves[c_idx]['y'])))
        return targets

    def on_slider_changed(self, val=None):
        if self._slider_base_state is None:
            return

        targets = self._get_target_points_map()
        scale_val = self.slider_scale.val
        smooth_val = int(self.slider_smooth.val)
        noise_coeff = self.slider_noise.val

        for c_idx, p_idxs in targets.items():
            orig_y = self._slider_base_state[c_idx]['y'].copy()
            p_idxs = np.array(p_idxs)
            if len(p_idxs) == 0:
                continue
            
            if scale_val != 1.0:
                orig_y[p_idxs] = orig_y[p_idxs] * scale_val

            if smooth_val > 1:
                smoothed = uniform_filter1d(orig_y.astype(float), size=smooth_val)
                orig_y[p_idxs] = smoothed[p_idxs]

            if noise_coeff > 0:
                data_range = np.std(orig_y[p_idxs]) if len(p_idxs) > 1 else np.mean(np.abs(orig_y[p_idxs]))
                if data_range == 0: data_range = 1.0

                rng = np.random.RandomState(self._noise_seed or 0)
                noise = rng.normal(0, data_range * 0.05 * noise_coeff, size=len(p_idxs))
                orig_y[p_idxs] += noise

            self.curves[c_idx]['y'][:] = orig_y
            self._update_spline(c_idx)

        self._expand_axes_if_needed()
        self._refresh_visual_style()
        self._update_selection_info()


# ===================== 演示数据 =====================
def _make_demo_curves(n_pts=80):
    """3 条干净利落的正弦曲线：低/中/高频，适合截图做 logo"""
    x = np.linspace(0, 3 * np.pi, n_pts)
    curves = [
        1.0  * np.sin(x),               # 基准波
        0.65 * np.sin(2.0 * x - 0.8),   # 倍频 + 小振幅
        0.45 * np.sin(3.5 * x + 0.5),   # 高频 + 更小振幅
    ]
    return np.column_stack(curves)  # (n_pts, 3)


# ===================== UI 工厂 =====================
def _build_ui():
    """创建 figure 和所有 axes，返回 (fig, axes_dict)"""
    # ── 现代柔和配色 ──
    BG_FIGURE  = '#eceff4'   # 整体背景
    BG_PLOT    = '#ffffff'   # 主绘图区
    GRID_COLOR = '#d8dee9'   # 网格线
    BG_SLIDER  = '#e5e9f0'   # 滑杆轨道底色
    BG_RADIO   = '#fafbfc'   # 单选面板
    BG_INPUT   = '#f4f5f8'   # 输入框底色

    fig = plt.figure(figsize=(16, 9))
    fig.patch.set_facecolor(BG_FIGURE)
    fig.canvas.manager.set_window_title("CurveMagician")

    ax = {
        'main':     plt.axes([0.05, 0.10, 0.70, 0.80], facecolor=BG_PLOT),
        # 左侧操作按钮组：Open | Save | Undo | Fit
        'open':     plt.axes([0.05, 0.94, 0.08, 0.035]),
        'save':     plt.axes([0.14, 0.94, 0.08, 0.035]),
        'undo':     plt.axes([0.23, 0.94, 0.08, 0.035]),
        'fit':      plt.axes([0.32, 0.94, 0.06, 0.035]),
        # 右侧参数组：+/-N | Points | Resample
        'neighbor': plt.axes([0.52, 0.94, 0.08, 0.035], facecolor=BG_INPUT),
        'npts':     plt.axes([0.61, 0.94, 0.07, 0.035], facecolor=BG_INPUT),
        'apply':    plt.axes([0.69, 0.94, 0.08, 0.035]),
        'scale':    plt.axes([0.80, 0.89, 0.18, 0.040], facecolor=BG_SLIDER),
        'smooth':   plt.axes([0.80, 0.84, 0.18, 0.040], facecolor=BG_SLIDER),
        'noise':    plt.axes([0.80, 0.79, 0.18, 0.040], facecolor=BG_SLIDER),
        'radio':    plt.axes([0.78, 0.08, 0.20, 0.68], facecolor=BG_RADIO),
    }

    # ── 主绘图区美化 ──
    main_ax = ax['main']
    main_ax.grid(True, linestyle='-', alpha=0.25, color=GRID_COLOR, linewidth=0.6)
    main_ax.tick_params(labelsize=9, colors='#4c566a')
    main_ax.spines['top'].set_visible(False)
    main_ax.spines['right'].set_visible(False)
    main_ax.spines['left'].set_color('#d8dee9')
    main_ax.spines['bottom'].set_color('#d8dee9')

    # ── 其他小 axes 统一隐藏边框 ──
    for key in ['open', 'save', 'undo', 'fit', 'apply', 'neighbor', 'npts']:
        _hide_axes_spines(ax[key])

    return fig, ax


def _hide_axes_spines(axes):
    """隐藏 axes 的四个边框和刻度"""
    for spine in axes.spines.values():
        spine.set_visible(False)
    axes.set_xticks([])
    axes.set_yticks([])


def _wire_widgets(editor, ax):
    """创建按钮/滑杆/输入框并连接到编辑器"""
    # ── 按钮配色（Nord 风格） ──
    C_OPEN   = '#5e81ac'  # 钢蓝
    C_SAVE   = '#a3be8c'  # 鼠尾草绿
    C_UNDO   = '#d08770'  # 暖橙
    C_FIT    = '#b48ead'  # 淡紫
    C_APPLY  = '#bf616a'  # 珊瑚红
    C_OPEN_H   = '#4c6f99'
    C_SAVE_H   = '#8caa72'
    C_UNDO_H   = '#c0745d'
    C_FIT_H    = '#a07a99'
    C_APPLY_H  = '#ad5058'

    # 左侧操作按钮组
    Button(ax['open'],  'Open',  color=C_OPEN,  hovercolor=C_OPEN_H).on_clicked(editor.on_open_clicked)
    Button(ax['save'],  'Save',  color=C_SAVE,  hovercolor=C_SAVE_H).on_clicked(editor.save_file)
    Button(ax['undo'],  'Undo',  color=C_UNDO,  hovercolor=C_UNDO_H).on_clicked(editor.undo)
    Button(ax['fit'],   'Fit',   color=C_FIT,   hovercolor=C_FIT_H).on_clicked(editor._auto_range_axes)

    # 右侧参数组
    TextBox(ax['neighbor'], "+/-N", initial="0").on_submit(editor.set_neighbor_num)

    resample_box = TextBox(ax['npts'], "", initial="")
    editor._resample_box = resample_box
    editor._typing_axes.add(ax['npts'])
    Button(ax['apply'], 'Resample', color=C_APPLY, hovercolor=C_APPLY_H).on_clicked(editor._do_resample)

    # ── 滑杆 ──
    SLIDER_COLOR = '#5e81ac'
    sliders = {
        'scale':  Slider(ax['scale'],  'Scale ',  0.0, 2.0, valinit=1.0, valfmt='%.2f', color=SLIDER_COLOR),
        'smooth': Slider(ax['smooth'], 'Smooth',  1,   15,  valinit=1,   valfmt='%d',   color=SLIDER_COLOR),
        'noise':  Slider(ax['noise'],  'Noise ',  0.0, 5.0, valinit=0.0, valfmt='%.1f', color=SLIDER_COLOR),
    }
    for s in sliders.values():
        s.label.set_size(10)
        s.label.set_color('#4c566a')
        s.label.set_fontweight('bold')
        s.valtext.set_fontsize(9)
        s.valtext.set_color('#4c566a')
        # 滑杆手柄（poly）美化
        if hasattr(s, 'poly'):
            s.poly.set_facecolor('#81a1c1')
            s.poly.set_edgecolor('#5e81ac')
            s.poly.set_linewidth(1.2)

    editor.slider_scale  = sliders['scale']
    editor.slider_smooth = sliders['smooth']
    editor.slider_noise  = sliders['noise']
    for s in sliders.values():
        s.on_changed(editor.on_slider_changed)


def _load_demo(editor, colors):
    """加载默认演示曲线"""
    data = _make_demo_curves()
    for idx in range(data.shape[1]):
        editor.add_curve(np.arange(data.shape[0]), data[:, idx],
                         color=colors[idx % len(colors)], name=f"Curve {idx}")
    editor._build_radio_panel(
        [colors[i % len(colors)] for i in range(data.shape[1])])
    editor._auto_range_axes()
    editor._save_current_state()
    editor._reset_slider_widgets()
    editor._update_selection_info()


# ===================== 主程序 =====================
if __name__ == "__main__":
    COLORS = ['#5e81ac', '#d08770', '#a3be8c', '#b48ead', '#bf616a',
              '#ebcb8b', '#81a1c1', '#8fbcbb', '#c9826b', '#7e9e6d',
              '#6a89cc', '#e55039', '#78e08f', '#6a89cc', '#b8e994']

    fig, ax = _build_ui()
    editor = HarmoniousCurvesEditor(
        ax['main'], fig, ax['neighbor'], ax_radio=ax['radio'], color_list=COLORS)

    _wire_widgets(editor, ax)
    _load_demo(editor, COLORS)
    editor.setup_drag_and_drop()
    plt.show()